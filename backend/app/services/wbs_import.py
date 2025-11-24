from __future__ import annotations

"""
Servizio di importazione/gestione WBS.

La gerarchia si divide in:
* WBS spaziale (livelli 1‑5) -> descrive lotto/edificio, livelli e componenti fisici.
* WBS6 (chiave analitica principale) -> aggrega sempre le quantità economiche e fa da
  pivot per confronti progetto/offerte.
* WBS7 (raggruppatori opzionali) -> sottocategorie della WBS6 per EPU o famiglie
  specifiche; tutte le offerte e i ritorni si agganciano a WBS6 e, se presente,
  al relativo WBS7 per round/impresa.
"""

from dataclasses import dataclass
from io import BytesIO
from datetime import datetime
from difflib import SequenceMatcher
import re
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import load_workbook
from sqlalchemy import update
from sqlmodel import Session, select

from app.db.models import Commessa
from app.db.models_wbs import Wbs6, Wbs7, WbsSpaziale
from app.db.models_wbs import Voce, VoceOfferta, VoceProgetto


WBS6_CODE_RE = re.compile(r"^([A-Za-z])\s*?(\d{3})$")
WBS7_CODE_RE = re.compile(r"^([A-Za-z])\s*?(\d{3})[.\s_-]?(\d{3})$")

HEADER_ALIASES = {
    "wbs 1 - lotto/edificio": 1,
    "wbs 2 - livelli": 2,
    "wbs 3 - ambiti omogenei": 3,
    "wbs 4 - appalto": 4,
    "wbs 5 - elementi funzionali": 5,
    "wbs 6 - categorie merceologiche": 6,
    "categorie merceologiche": 6,
    "wbs6": 6,
    "wbs 6": 6,
    "raggruppatore epu": 7,
    "wbs 7": 7,
}

HEADER_MARKERS = {"codice", "descrizione"}


@dataclass
class ParsedSpatialLevel:
    level: int
    code: str
    description: Optional[str]


@dataclass
class ParsedWbsRow:
    spatial_levels: list[ParsedSpatialLevel]
    wbs6_code: str
    wbs6_description: str
    wbs7_code: Optional[str]
    wbs7_description: Optional[str]


@dataclass
class WbsImportStats:
    rows_total: int = 0
    spaziali_inserted: int = 0
    spaziali_updated: int = 0
    wbs6_inserted: int = 0
    wbs6_updated: int = 0
    wbs7_inserted: int = 0
    wbs7_updated: int = 0


@dataclass
class Wbs6NormalizationMatch:
    current_code: str
    current_description: Optional[str]
    suggested_code: Optional[str]
    suggested_description: Optional[str]
    score: float
    reason: str


@dataclass
class Wbs7NormalizationMatch:
    current_code: Optional[str]
    current_description: Optional[str]
    wbs6_code: Optional[str]
    suggested_code: Optional[str]
    suggested_description: Optional[str]
    suggested_wbs6_code: Optional[str]
    score: float
    reason: str


@dataclass
class WbsNormalizationResult:
    wbs6: list[Wbs6NormalizationMatch]
    wbs7: list[Wbs7NormalizationMatch]


class WbsImportService:
    """Parser e persistenza per file Excel WBS."""

    _UNSET = object()

    @classmethod
    def import_from_upload(
        cls,
        session: Session,
        commessa: Commessa,
        *,
        file_bytes: bytes,
        mode: str = "create",
    ) -> WbsImportStats:
        if mode not in {"create", "update"}:
            raise ValueError("Modalità non supportata, usa create oppure update")

        has_existing = session.exec(
            select(Wbs6).where(Wbs6.commessa_id == commessa.id)
        ).first()
        if has_existing and mode == "create":
            raise ValueError(
                "La commessa ha già una WBS importata. Usa PUT per effettuare un aggiornamento."
            )

        rows = cls._parse_excel(BytesIO(file_bytes))
        if not rows:
            raise ValueError("Il file non contiene righe WBS valide")

        context = _WbsPersistenceContext(session, commessa.id)
        if session.in_transaction():
            return context.persist(rows)
        with session.begin():
            return context.persist(rows)

    @staticmethod
    def fetch_commessa_wbs(
        session: Session, commessa_id: int
    ) -> tuple[list[WbsSpaziale], list[Wbs6], list[Wbs7]]:
        spaziali = session.exec(
            select(WbsSpaziale).where(WbsSpaziale.commessa_id == commessa_id)
        ).all()
        wbs6_nodes = session.exec(
            select(Wbs6).where(Wbs6.commessa_id == commessa_id)
        ).all()
        wbs7_nodes = session.exec(
            select(Wbs7).where(Wbs7.commessa_id == commessa_id)
        ).all()
        return list(spaziali), list(wbs6_nodes), list(wbs7_nodes)

    @classmethod
    def build_normalization_plan_from_excel(
        cls,
        session: Session,
        commessa: Commessa,
        *,
        file_bytes: bytes,
        min_score_wbs6: float = 0.72,
        min_score_wbs7: float = 0.68,
    ) -> WbsNormalizationResult:
        """
        Genera un piano di normalizzazione WBS a partire da un Excel di riferimento.

        Per ogni WBS6/WBS7 già presente in commessa propone un codice target basato
        su una corrispondenza per descrizione con i nodi definiti nel file fornito.
        Non esegue modifiche a DB: restituisce solo i suggerimenti.
        """
        rows = cls._parse_excel(BytesIO(file_bytes))
        if not rows:
            raise ValueError("Il file non contiene righe WBS valide")

        reference_wbs6: Dict[str, dict] = {}
        reference_wbs6_index: list[dict] = []
        reference_wbs7: Dict[str, list[dict]] = {}

        for row in rows:
            norm_desc = cls._normalize_match_text(row.wbs6_description)
            reference_wbs6[row.wbs6_code] = {
                "code": row.wbs6_code,
                "description": row.wbs6_description,
                "norm_desc": norm_desc,
            }
            reference_wbs6_index.append(reference_wbs6[row.wbs6_code])
            if row.wbs7_code:
                bucket = reference_wbs7.setdefault(row.wbs6_code, [])
                bucket.append(
                    {
                        "code": row.wbs7_code,
                        "description": row.wbs7_description,
                        "norm_desc": cls._normalize_match_text(row.wbs7_description),
                    }
                )

        existing_wbs6 = session.exec(
            select(Wbs6).where(Wbs6.commessa_id == commessa.id)
        ).all()
        wbs6_matches: list[Wbs6NormalizationMatch] = []
        wbs6_target_by_id: Dict[int, str] = {}

        for node in existing_wbs6:
            direct_match = reference_wbs6.get(node.code)
            if direct_match:
                wbs6_target_by_id[node.id] = node.code
                wbs6_matches.append(
                    Wbs6NormalizationMatch(
                        current_code=node.code,
                        current_description=node.description,
                        suggested_code=node.code,
                        suggested_description=direct_match["description"],
                        score=1.0,
                        reason="code_match",
                    )
                )
                continue

            best = cls._pick_best_by_description(
                node.description or node.label,
                reference_wbs6_index,
            )
            if best and best[1] >= min_score_wbs6:
                target = best[0]
                wbs6_target_by_id[node.id] = target["code"]
                wbs6_matches.append(
                    Wbs6NormalizationMatch(
                        current_code=node.code,
                        current_description=node.description,
                        suggested_code=target["code"],
                        suggested_description=target["description"],
                        score=best[1],
                        reason="description_match",
                    )
                )
            else:
                wbs6_target_by_id[node.id] = node.code
                wbs6_matches.append(
                    Wbs6NormalizationMatch(
                        current_code=node.code,
                        current_description=node.description,
                        suggested_code=None,
                        suggested_description=None,
                        score=0.0,
                        reason="unmatched",
                    )
                )

        existing_wbs6_map: Dict[int, Wbs6] = {node.id: node for node in existing_wbs6}
        existing_wbs7 = session.exec(
            select(Wbs7).where(Wbs7.commessa_id == commessa.id)
        ).all()
        wbs7_matches: list[Wbs7NormalizationMatch] = []

        for node in existing_wbs7:
            parent = existing_wbs6_map.get(node.wbs6_id)
            parent_code = parent.code if parent else None
            target_parent_code = wbs6_target_by_id.get(node.wbs6_id, parent_code)
            candidates = reference_wbs7.get(target_parent_code or "", [])
            direct_candidate = None
            if candidates and node.code:
                direct_candidate = next(
                    (cand for cand in candidates if cand["code"] == node.code),
                    None,
                )
            if direct_candidate:
                wbs7_matches.append(
                    Wbs7NormalizationMatch(
                        current_code=node.code,
                        current_description=node.description,
                        wbs6_code=parent_code,
                        suggested_code=direct_candidate["code"],
                        suggested_description=direct_candidate["description"],
                        suggested_wbs6_code=target_parent_code,
                        score=1.0,
                        reason="code_match",
                    )
                )
                continue

            best7 = cls._pick_best_by_description(
                node.description,
                candidates,
            )
            if best7 and best7[1] >= min_score_wbs7:
                target = best7[0]
                wbs7_matches.append(
                    Wbs7NormalizationMatch(
                        current_code=node.code,
                        current_description=node.description,
                        wbs6_code=parent_code,
                        suggested_code=target["code"],
                        suggested_description=target["description"],
                        suggested_wbs6_code=target_parent_code,
                        score=best7[1],
                        reason="description_match",
                    )
                )
            else:
                wbs7_matches.append(
                    Wbs7NormalizationMatch(
                        current_code=node.code,
                        current_description=node.description,
                        wbs6_code=parent_code,
                        suggested_code=None,
                        suggested_description=None,
                        suggested_wbs6_code=None,
                        score=0.0,
                        reason="unmatched",
                    )
                )

        return WbsNormalizationResult(wbs6=wbs6_matches, wbs7=wbs7_matches)

    @classmethod
    def update_spatial_node(
        cls,
        session: Session,
        commessa_id: int,
        node_id: int,
        *,
        code: object = _UNSET,
        description: object = _UNSET,
        importo_totale: object = _UNSET,
        parent_id: object = _UNSET,
        level: object = _UNSET,
    ) -> WbsSpaziale:
        node = session.get(WbsSpaziale, node_id)
        if not node or node.commessa_id != commessa_id:
            raise ValueError("Nodo WBS spaziale non trovato per la commessa")

        def _normalize_code(value: str) -> str:
            normalized = value.strip()
            if not normalized:
                raise ValueError("Il codice WBS deve contenere almeno un carattere")
            return normalized

        with session.begin():
            if code is not cls._UNSET:
                normalized = _normalize_code(str(code))
                duplicate = session.exec(
                    select(WbsSpaziale).where(
                        WbsSpaziale.commessa_id == commessa_id,
                        WbsSpaziale.level == node.level,
                        WbsSpaziale.code == normalized,
                        WbsSpaziale.id != node.id,
                    )
                ).first()
                if duplicate:
                    raise ValueError(
                        f"Esiste già un nodo di livello {node.level} con il codice {normalized}"
                    )
                node.code = normalized
            if description is not cls._UNSET:
                node.description = str(description).strip() if description else None
            if importo_totale is not cls._UNSET:
                node.importo_totale = (
                    float(importo_totale) if importo_totale is not None else None
                )
            if parent_id is not cls._UNSET:
                if parent_id is None:
                    node.parent_id = None
                else:
                    parent = session.get(WbsSpaziale, int(parent_id))
                    if not parent or parent.commessa_id != commessa_id:
                        raise ValueError("Nodo padre non valido per questa commessa")
                    if parent.id == node.id or parent.level >= node.level:
                        raise ValueError(
                            "Il nodo padre deve appartenere a un livello inferiore"
                        )
                    node.parent_id = parent.id
            if level is not cls._UNSET:
                if level is None or not (1 <= int(level) <= 5):
                    raise ValueError("Il livello WBS deve essere compreso tra 1 e 5")
                level_value = int(level)
                if node.level != level_value:
                    node.level = level_value
            session.add(node)
        session.refresh(node)
        return node

    @classmethod
    def update_wbs6_node(
        cls,
        session: Session,
        commessa_id: int,
        node_id: int,
        *,
        code: object = _UNSET,
        description: object = _UNSET,
        wbs_spaziale_id: object = _UNSET,
    ) -> Wbs6:
        node = session.get(Wbs6, node_id)
        if not node or node.commessa_id != commessa_id:
            raise ValueError("Nodo WBS6 non trovato per la commessa")

        with session.begin():
            if code is not cls._UNSET:
                normalized = cls._validate_wbs6_code(str(code))
                duplicate = session.exec(
                    select(Wbs6).where(
                        Wbs6.commessa_id == commessa_id,
                        Wbs6.code == normalized,
                        Wbs6.id != node.id,
                    )
                ).first()
                if duplicate:
                    raise ValueError(f"Il codice WBS6 {normalized} è già utilizzato")
                node.code = normalized
            if description is not cls._UNSET:
                text = str(description).strip()
                if not text:
                    raise ValueError("La descrizione WBS6 non può essere vuota")
                node.description = text
            if wbs_spaziale_id is not cls._UNSET:
                if wbs_spaziale_id is None:
                    node.wbs_spaziale_id = None
                else:
                    target = session.get(WbsSpaziale, int(wbs_spaziale_id))
                    if not target or target.commessa_id != commessa_id:
                        raise ValueError("Nodo spaziale associato non valido")
                    node.wbs_spaziale_id = target.id
            node.label = f"{node.code} - {node.description}"
            session.add(node)
            cls._touch_related_by_wbs6(session, node.id)
        session.refresh(node)
        return node

    @classmethod
    def update_wbs7_node(
        cls,
        session: Session,
        commessa_id: int,
        node_id: int,
        *,
        code: object = _UNSET,
        description: object = _UNSET,
    ) -> Wbs7:
        node = session.get(Wbs7, node_id)
        if not node or node.commessa_id != commessa_id:
            raise ValueError("Nodo WBS7 non trovato per la commessa")

        with session.begin():
            if code is not cls._UNSET:
                normalized = cls._validate_wbs7_code(str(code))
                duplicate = session.exec(
                    select(Wbs7).where(
                        Wbs7.commessa_id == commessa_id,
                        Wbs7.wbs6_id == node.wbs6_id,
                        Wbs7.code == normalized,
                        Wbs7.id != node.id,
                    )
                ).first()
                if duplicate:
                    raise ValueError("Questo codice WBS7 è già presente per la WBS6 selezionata")
                node.code = normalized
            if description is not cls._UNSET:
                node.description = str(description).strip() if description else None
            session.add(node)
            cls._touch_related_by_wbs7(session, node.id)
        session.refresh(node)
        return node

    @classmethod
    def _parse_excel(cls, handle: BytesIO) -> list[ParsedWbsRow]:
        workbook = load_workbook(handle, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        header_row, columns = cls._detect_columns(rows)
        data_start = header_row + 2

        memory = {
            level: {"code": None, "description": None}
            for level in range(1, 7)
        }

        parsed: list[ParsedWbsRow] = []
        for idx in range(data_start, len(rows)):
            row = rows[idx]
            if cls._is_header_like(row):
                continue
            spatial_levels = cls._extract_spatial_levels(row, columns, memory)
            wbs6_code, wbs6_desc = cls._extract_wbs6(row, columns, memory)
            if not wbs6_code:
                continue
            wbs7_code, wbs7_desc = cls._extract_wbs7(row, columns)
            parsed.append(
                ParsedWbsRow(
                    spatial_levels=spatial_levels,
                    wbs6_code=wbs6_code,
                    wbs6_description=wbs6_desc,
                    wbs7_code=wbs7_code,
                    wbs7_description=wbs7_desc,
                )
            )
        return parsed

    @staticmethod
    def _is_header_like(row: Sequence[object | None]) -> bool:
        values = {
            (str(cell).strip().lower())
            for cell in row
            if cell is not None and str(cell).strip()
        }
        return bool(values) and values.issubset(HEADER_MARKERS)

    @classmethod
    def _detect_columns(
        cls, rows: list[Tuple[object, ...]]
    ) -> tuple[int, Dict[int, tuple[int, int]]]:
        columns: Dict[int, tuple[int, int]] = {}
        header_row = -1
        for idx, row in enumerate(rows):
            for col_idx, cell in enumerate(row):
                if not cell:
                    continue
                normalized = str(cell).strip().lower()
                level = HEADER_ALIASES.get(normalized)
                if level:
                    columns[level] = (col_idx, col_idx + 1)
            if 6 in columns:
                header_row = idx
                break
        if header_row < 0 or 6 not in columns:
            raise ValueError("Intestazioni WBS non riconosciute nel file")
        return header_row, columns

    @classmethod
    def _extract_spatial_levels(
        cls,
        row: Sequence[object | None],
        columns: Dict[int, tuple[int, int]],
        memory: Dict[int, Dict[str, Optional[str]]],
    ) -> list[ParsedSpatialLevel]:
        result: list[ParsedSpatialLevel] = []
        for level in range(1, 6):
            col = columns.get(level)
            if not col:
                continue
            code_raw = cls._normalize_code(row, col[0])
            desc_raw = cls._normalize_text(row, col[1])
            if code_raw:
                memory[level]["code"] = code_raw
                if desc_raw:
                    memory[level]["description"] = desc_raw
                cls._reset_lower_levels(memory, level)
            elif desc_raw:
                memory[level]["description"] = desc_raw
            code = memory[level]["code"]
            if code:
                result.append(
                    ParsedSpatialLevel(
                        level=level,
                        code=code,
                        description=memory[level]["description"],
                    )
                )
        return result

    @classmethod
    def _extract_wbs6(
        cls,
        row: Sequence[object | None],
        columns: Dict[int, tuple[int, int]],
        memory: Dict[int, Dict[str, Optional[str]]],
    ) -> tuple[Optional[str], Optional[str]]:
        col = columns.get(6)
        if not col:
            return None, None
        code = cls._normalize_wbs6(row[col[0]])
        desc = cls._normalize_text(row, col[1])
        if code:
            memory[6]["code"] = code
        if desc:
            memory[6]["description"] = desc
        code_value = memory[6]["code"]
        desc_value = memory[6]["description"]
        if code_value and not desc_value:
            desc_value = f"WBS6 {code_value}"
        return code_value, desc_value

    @classmethod
    def _extract_wbs7(
        cls,
        row: Sequence[object | None],
        columns: Dict[int, tuple[int, int]],
    ) -> tuple[Optional[str], Optional[str]]:
        col = columns.get(7)
        if not col:
            return None, None
        code = cls._normalize_wbs7(row[col[0]])
        desc = cls._normalize_text(row, col[1])
        return code, desc

    @staticmethod
    def _reset_lower_levels(
        memory: Dict[int, Dict[str, Optional[str]]], level: int
    ) -> None:
        for deeper in range(level + 1, 6):
            memory[deeper]["code"] = None
            memory[deeper]["description"] = None

    @staticmethod
    def _normalize_code(row: Sequence[object | None], index: int) -> Optional[str]:
        if index >= len(row):
            return None
        value = row[index]
        if value is None:
            return None
        text = str(value).strip()
        lowered = text.lower()
        if lowered in HEADER_MARKERS:
            return None
        return text or None

    @staticmethod
    def _normalize_text(row: Sequence[object | None], index: int) -> Optional[str]:
        if index >= len(row):
            return None
        value = row[index]
        if value is None:
            return None
        text = str(value).replace("\n", " ").strip()
        if text.lower() in HEADER_MARKERS:
            return None
        return text or None

    @staticmethod
    def _normalize_wbs6(value: object | None) -> Optional[str]:
        if value is None:
            return None
        text = str(value).strip()
        match = WBS6_CODE_RE.match(text.replace(" ", ""))
        if not match:
            return None
        letter, digits = match.groups()
        return f"{letter.upper()}{digits}"

    @staticmethod
    def _normalize_wbs7(value: object | None) -> Optional[str]:
        if value is None:
            return None
        text = str(value).strip()
        match = WBS7_CODE_RE.match(text.replace(" ", ""))
        if not match:
            return None
        letter, digits, suffix = match.groups()
        return f"{letter.upper()}{digits}.{suffix}"

    @staticmethod
    def _validate_wbs6_code(value: str) -> str:
        normalized = value.strip()
        match = WBS6_CODE_RE.match(normalized.replace(" ", ""))
        if not match:
            raise ValueError("Formato WBS6 non valido (atteso A###)")
        letter, digits = match.groups()
        return f"{letter.upper()}{digits}"

    @staticmethod
    def _validate_wbs7_code(value: str) -> str:
        normalized = value.strip()
        match = WBS7_CODE_RE.match(normalized.replace(" ", ""))
        if not match:
            raise ValueError("Formato WBS7 non valido (atteso A###.###)")
        letter, digits, suffix = match.groups()
        return f"{letter.upper()}{digits}.{suffix}"

    @staticmethod
    def _touch_related_by_wbs6(session: Session, wbs6_id: int) -> None:
        voce_ids = session.exec(
            select(Voce.id).where(Voce.wbs6_id == wbs6_id)
        ).all()
        WbsImportService._touch_related_tables(session, [vid for vid in voce_ids if vid is not None])

    @staticmethod
    def _touch_related_by_wbs7(session: Session, wbs7_id: int) -> None:
        voce_ids = session.exec(
            select(Voce.id).where(Voce.wbs7_id == wbs7_id)
        ).all()
        WbsImportService._touch_related_tables(session, [vid for vid in voce_ids if vid is not None])

    @staticmethod
    def _touch_related_tables(session: Session, voce_ids: list[int]) -> None:
        if not voce_ids:
            return
        now = datetime.utcnow()
        session.exec(
            update(Voce)
            .where(Voce.id.in_(voce_ids))
            .values(updated_at=now)
        )
        session.exec(
            update(VoceProgetto)
            .where(VoceProgetto.voce_id.in_(voce_ids))
            .values(updated_at=now)
        )
        session.exec(
            update(VoceOfferta)
            .where(VoceOfferta.voce_id.in_(voce_ids))
            .values(updated_at=now)
        )

    @classmethod
    def _pick_best_by_description(
        cls,
        description: Optional[str],
        candidates: Sequence[dict],
    ) -> Optional[tuple[dict, float]]:
        needle = cls._normalize_match_text(description)
        if not needle or not candidates:
            return None
        best_score = 0.0
        best_candidate: Optional[dict] = None
        for candidate in candidates:
            score = cls._similarity_score(needle, candidate.get("norm_desc", ""))
            if score > best_score:
                best_score = score
                best_candidate = candidate
        if best_candidate is None:
            return None
        return best_candidate, best_score

    @staticmethod
    def _normalize_match_text(value: Optional[str]) -> str:
        if not value:
            return ""
        cleaned = re.sub(r"[^a-z0-9]+", " ", value.lower())
        cleaned = re.sub(r"\s+", " ", cleaned).strip()
        return cleaned

    @staticmethod
    def _similarity_score(left: str, right: str) -> float:
        if not left or not right:
            return 0.0
        ratio = SequenceMatcher(None, left, right).ratio()
        left_tokens = set(left.split())
        right_tokens = set(right.split())
        overlap = (
            len(left_tokens & right_tokens) / max(len(left_tokens), len(right_tokens))
            if left_tokens and right_tokens
            else 0.0
        )
        return max(ratio, overlap)


class _WbsPersistenceContext:
    """Upsert atomico dei nodi WBS con statistiche di import."""

    def __init__(self, session: Session, commessa_id: int) -> None:
        self.session = session
        self.commessa_id = commessa_id
        self.spatial_cache: Dict[tuple[int, str], WbsSpaziale] = {}
        self.wbs6_cache: Dict[str, Wbs6] = {}
        self.wbs7_cache: Dict[tuple[int, str], Wbs7] = {}

    def persist(self, rows: Iterable[ParsedWbsRow]) -> WbsImportStats:
        stats = WbsImportStats()
        for row in rows:
            stats.rows_total += 1
            leaf = self._ensure_spatial_levels(row.spatial_levels, stats)
            wbs6 = self._upsert_wbs6(row, leaf, stats)
            self._upsert_wbs7(row, wbs6, stats)
        return stats

    def _ensure_spatial_levels(
        self,
        levels: Sequence[ParsedSpatialLevel],
        stats: WbsImportStats,
    ) -> Optional[WbsSpaziale]:
        parent: Optional[WbsSpaziale] = None
        for level in levels:
            node, created, updated = self._upsert_spatial(level, parent)
            if created:
                stats.spaziali_inserted += 1
            elif updated:
                stats.spaziali_updated += 1
            parent = node
        return parent

    def _upsert_spatial(
        self,
        level: ParsedSpatialLevel,
        parent: Optional[WbsSpaziale],
    ) -> tuple[WbsSpaziale, bool, bool]:
        key = (level.level, level.code)
        node = self.spatial_cache.get(key)
        created = False
        updated = False
        if not node:
            node = self.session.exec(
                select(WbsSpaziale).where(
                    WbsSpaziale.commessa_id == self.commessa_id,
                    WbsSpaziale.level == level.level,
                    WbsSpaziale.code == level.code,
                )
            ).first()
        if not node:
            node = WbsSpaziale(
                commessa_id=self.commessa_id,
                parent_id=parent.id if parent else None,
                level=level.level,
                code=level.code,
                description=level.description,
            )
            self.session.add(node)
            self.session.flush()
            created = True
        else:
            if level.description and node.description != level.description:
                node.description = level.description
                updated = True
            new_parent_id = parent.id if parent else None
            if node.parent_id != new_parent_id:
                node.parent_id = new_parent_id
                updated = True
            if updated:
                self.session.add(node)
        self.spatial_cache[key] = node
        return node, created, updated

    def _upsert_wbs6(
        self,
        row: ParsedWbsRow,
        leaf: Optional[WbsSpaziale],
        stats: WbsImportStats,
    ) -> Wbs6:
        node = self.wbs6_cache.get(row.wbs6_code)
        created = False
        if not node:
            node = self.session.exec(
                select(Wbs6).where(
                    Wbs6.commessa_id == self.commessa_id,
                    Wbs6.code == row.wbs6_code,
                )
            ).first()
        if not node:
            node = Wbs6(
                commessa_id=self.commessa_id,
                wbs_spaziale_id=leaf.id if leaf else None,
                code=row.wbs6_code,
                description=row.wbs6_description,
                label=f"{row.wbs6_code} - {row.wbs6_description}",
            )
            self.session.add(node)
            self.session.flush()
            created = True
            stats.wbs6_inserted += 1
        else:
            updated = False
            if leaf and node.wbs_spaziale_id != leaf.id:
                node.wbs_spaziale_id = leaf.id
                updated = True
            if node.description != row.wbs6_description:
                node.description = row.wbs6_description
                node.label = f"{row.wbs6_code} - {row.wbs6_description}"
                updated = True
            if updated:
                stats.wbs6_updated += 1
                self.session.add(node)
        self.wbs6_cache[row.wbs6_code] = node
        return node

    def _upsert_wbs7(
        self,
        row: ParsedWbsRow,
        wbs6: Wbs6,
        stats: WbsImportStats,
    ) -> None:
        if not row.wbs7_code:
            return
        key = (wbs6.id, row.wbs7_code)
        node = self.wbs7_cache.get(key)
        if not node:
            node = self.session.exec(
                select(Wbs7).where(
                    Wbs7.commessa_id == self.commessa_id,
                    Wbs7.wbs6_id == wbs6.id,
                    Wbs7.code == row.wbs7_code,
                )
            ).first()
        if not node:
            node = Wbs7(
                commessa_id=self.commessa_id,
                wbs6_id=wbs6.id,
                code=row.wbs7_code,
                description=row.wbs7_description,
            )
            self.session.add(node)
            self.session.flush()
            stats.wbs7_inserted += 1
        else:
            if row.wbs7_description and node.description != row.wbs7_description:
                node.description = row.wbs7_description
                stats.wbs7_updated += 1
                self.session.add(node)
        self.wbs7_cache[key] = node
