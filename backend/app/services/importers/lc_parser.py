from __future__ import annotations

from pathlib import Path
from typing import Sequence

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from app.excel import ParsedComputo, ParsedVoce, ParsedWbsLevel
from app.services.importers.parser import (
    _CustomReturnParseResult,
    _apply_column_filter,
    _cell_to_float,
    _columns_to_indexes,
    _combine_code,
    _combine_text,
    _detect_column_suggestions,
    _drop_empty_columns,
    _has_external_formula,
    _locate_header_row,
    _looks_like_wbs7_code,
    _normalize_wbs7_code,
    _sanitize_price_candidate,
    _select_sheet,
    _single_column_index,
)
from app.services.importers.common import _calculate_line_amount, _ceil_amount


def parse_lc_return_excel(
    file_path: Path,
    sheet_name: str | None,
    code_columns: Sequence[str],
    description_columns: Sequence[str],
    price_column: str,
    quantity_column: str | None = None,
    progressive_column: str | None = None,
) -> _CustomReturnParseResult:
    """
    Parser LC lineare: una riga = una voce, senza combinare header/totali.
    """
    workbook = load_workbook(filename=file_path, data_only=True, read_only=True)
    try:
        sheet = _select_sheet(workbook, sheet_name)
        raw_rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    rows, dropped_columns, kept_column_indexes = _drop_empty_columns(raw_rows)
    header_idx = _locate_header_row(rows)
    if header_idx is None:
        raise ValueError("Il foglio Excel selezionato non contiene righe valide da importare")
    data_rows = rows[header_idx + 1 :]
    if not data_rows:
        raise ValueError("Il foglio Excel selezionato non contiene dati dopo l'intestazione")

    header_row = rows[header_idx]
    detection = _detect_column_suggestions(rows, header_idx)
    suggestions = detection.get("suggestions") if detection else {}
    column_warnings: list[str] = []
    if dropped_columns:
        column_warnings.append(f"Ignorate automaticamente {dropped_columns} colonne completamente vuote.")

    code_indexes = _ensure_indexes_lc("codice", code_columns, data_rows, header_row, column_warnings, suggestions)
    description_indexes = _ensure_indexes_lc("descrizione", description_columns, data_rows, header_row, column_warnings, suggestions)

    price_index = _single_column_index(price_column, "prezzo unitario", header_row=header_row)
    if not _has_values(data_rows, [price_index]):
        raise ValueError(
            f"La colonna prezzo selezionata ({get_column_letter(price_index + 1)}) non contiene valori. "
            "Verifica la colonna scelta nel foglio."
        )

    quantity_index = _single_column_index(quantity_column, "quantita", header_row=header_row, required=False)
    progressive_index = _single_column_index(progressive_column, "progressivo", header_row=header_row, required=False)

    workbook_formulas = load_workbook(filename=file_path, data_only=False, read_only=True)
    try:
        formula_sheet = _select_sheet(workbook_formulas, sheet_name)
        raw_formula_rows = list(formula_sheet.iter_rows(min_row=header_idx + 2, values_only=False))
        formula_rows = _apply_column_filter(raw_formula_rows, kept_column_indexes)
    finally:
        workbook_formulas.close()

    formula_rows_iter = iter(formula_rows)
    voci: list[ParsedVoce] = []
    ordine = 0
    for row in data_rows:
        formula_row = next(formula_rows_iter, ())
        if not _row_has_values(row):
            continue
        codice = _combine_code(row, code_indexes) or _combine_text(row, code_indexes)
        descrizione = _combine_text(row, description_indexes)
        raw_price = _cell_to_float(row, price_index)
        quantita = _cell_to_float(row, quantity_index) if quantity_index is not None else None
        progressivo_value = _cell_to_progressive(row, progressive_index)

        if raw_price is None:
            continue

        formula_cell = formula_row[price_index] if price_index < len(formula_row) else None
        if _has_external_formula(formula_cell):
            raise ValueError(
                "La colonna prezzo contiene formule collegate a file esterni. Apri il file in Excel e incolla i valori numerici prima di importare."
            )
        quantity_formula_cell = formula_row[quantity_index] if quantity_index is not None and quantity_index < len(formula_row) else None
        if quantity_index is not None and _has_external_formula(quantity_formula_cell):
            raise ValueError(
                "La colonna quantit�� contiene formule collegate a file esterni. Incolla i valori numerici prima dell'import."
            )

        prezzo_value = _sanitize_price_candidate(raw_price)
        if prezzo_value is None:
            continue
        quantita_value = quantita if quantita not in (None,) else None
        importo_value = None
        if quantita_value is not None:
            _, importo_value = _calculate_line_amount(quantita_value, round(prezzo_value, 4))
        else:
            importo_value = _ceil_amount(prezzo_value)

        wbs_levels: list[ParsedWbsLevel] = []
        normalized_code_value = _normalize_wbs7_code(codice)
        voce_descrizione = descrizione or codice or "Voce senza descrizione"
        if _looks_like_wbs7_code(normalized_code_value):
            wbs_levels.append(
                ParsedWbsLevel(
                    level=7,
                    code=normalized_code_value,
                    description=voce_descrizione,
                )
            )
        voci.append(
            ParsedVoce(
                ordine=ordine,
                progressivo=progressivo_value,
                codice=codice,
                descrizione=voce_descrizione,
                wbs_levels=wbs_levels,
                unita_misura=None,
                quantita=quantita_value,
                prezzo_unitario=round(prezzo_value, 4),
                importo=importo_value,
                note=None,
                metadata=None,
            )
        )
        ordine += 1

    computo = ParsedComputo(
        titolo=None,
        voci=voci,
        totale_quantita=None,
        totale_importo=None,
    )
    return _CustomReturnParseResult(computo=computo, column_warnings=column_warnings)


def _has_values(rows: list, indexes: Sequence[int]) -> bool:
    for row in rows:
        for idx in indexes:
            if idx < len(row) and _cell_has_content(row[idx]):
                return True
    return False


# Re-export internal helpers we rely on
from app.services.importers.parser import _cell_has_content, _cell_to_progressive, _row_has_values  # noqa: E402,F401


def _ensure_indexes_lc(
    name: str,
    columns: Sequence[str],
    data_rows,
    header_row,
    warnings: list[str],
    suggestions,
) -> list[int]:
    try:
        return _columns_to_indexes(columns, name, header_row=header_row, required=False)
    except ValueError:
        indexes: list[int] = []
    suggestion = suggestions.get(name) if suggestions else None
    if suggestion is not None:
        warnings.append(
            f"Colonna {name} non trovata nella configurazione salvata. "
            f"Utilizzo automaticamente '{suggestion.header_label or suggestion.column_letter}'."
        )
        return [suggestion.column_index]
    return indexes
