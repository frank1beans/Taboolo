from __future__ import annotations

import re
from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path
from typing import Any, Sequence

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

from app.excel import ParsedComputo, ParsedVoce, ParsedWbsLevel
from app.services.importers.common import (
    _calculate_line_amount,
    _ceil_amount,
    _normalize_wbs7_code,
    _looks_like_wbs7_code,
)


def _parse_custom_return_excel(
    file_path: Path,
    sheet_name: str | None,
    code_columns: Sequence[str],
    description_columns: Sequence[str],
    price_column: str,
    quantity_column: str | None = None,
    progressive_column: str | None = None,
    *,
    combine_totals: bool = True,
) -> ParsedComputo:
    workbook = load_workbook(filename=file_path, data_only=True, read_only=True)
    try:
        sheet = _select_sheet(workbook, sheet_name)
        raw_rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    # Pre-normalizza la tabella: elimina colonne completamente vuote e riempie i None con "" dove serve
    rows, dropped_columns, kept_column_indexes = _drop_empty_columns(raw_rows)
    header_idx = _locate_header_row(rows)
    if header_idx is None:
        raise ValueError("Il foglio Excel selezionato non contiene righe valide da importare")
    data_rows = rows[header_idx + 1 :]
    if not data_rows:
        raise ValueError("Il foglio Excel selezionato non contiene dati dopo l'intestazione")

    header_row = rows[header_idx]
    detection = _detect_column_suggestions(rows, header_idx)
    suggestions: dict[str, _ColumnSuggestion] = detection.get("suggestions") if detection else {}
    profiles: list[_ColumnProfile] = detection.get("profiles") if detection else []

    def _format_profiles() -> str:
        entries: list[str] = []
        for profile in profiles[:10]:
            label = profile.header_label or "�?"
            entries.append(f"{profile.letter}: {label}")
        return ", ".join(entries)

    def _warn_and_use(target: str, suggestion: _ColumnSuggestion | None, warnings: list[str]) -> int | None:
        if suggestion is None:
            return None
        label = suggestion.header_label or suggestion.column_letter
        warnings.append(
            f"Colonna {target} non trovata nella configurazione salvata. "
            f"Utilizzo automaticamente '{label}' ({suggestion.column_letter})."
        )
        return suggestion.column_index

    def _ensure_indexes(name: str, indexes: list[int], warnings: list[str]) -> list[int]:
        if indexes and _column_has_values(data_rows, indexes):
            return indexes
        suggestion = suggestions.get(name)
        fallback = _warn_and_use(name, suggestion, warnings)
        if fallback is not None:
            return [fallback]
        available = _format_profiles()
        raise ValueError(
            f"Impossibile individuare la colonna {name}. "
            f"Intestazioni rilevate: {available or 'nessuna intestazione valida.'}"
        )

    column_warnings: list[str] = []
    if dropped_columns:
        column_warnings.append(
            f"Ignorate automaticamente {dropped_columns} colonne completamente vuote."
        )
    try:
        code_indexes = _columns_to_indexes(code_columns, "codice", header_row=header_row, required=False)
    except ValueError:
        code_indexes = []
    try:
        description_indexes = _columns_to_indexes(description_columns, "descrizione", header_row=header_row, required=False)
    except ValueError:
        description_indexes = []
    try:
        price_index = _single_column_index(price_column, "prezzo unitario", header_row=header_row)
    except ValueError:
        price_index = None
    try:
        quantity_index = _single_column_index(quantity_column, "quantita", header_row=header_row, required=False)
    except ValueError:
        quantity_index = None
    try:
        progressive_index = _single_column_index(progressive_column, "progressivo", header_row=header_row, required=False)
    except ValueError:
        progressive_index = None

    code_indexes = _ensure_indexes("codice", code_indexes, column_warnings)
    description_indexes = _ensure_indexes("descrizione", description_indexes, column_warnings)
    if price_index is None:
        suggestion = suggestions.get("prezzo")
        price_index = _warn_and_use("prezzo", suggestion, column_warnings)
    if price_index is None:
        available = _format_profiles()
        raise ValueError(
            f"Impossibile individuare la colonna prezzo unitario. "
            f"Intestazioni rilevate: {available or 'nessuna intestazione valida.'}"
        )
    if not _column_has_values(data_rows, [price_index]):
        raise ValueError(
            f"La colonna prezzo selezionata ({get_column_letter(price_index + 1)}) non contiene valori. "
            "Verifica la colonna scelta nel foglio."
        )
    if quantity_index is None:
        quantity_index = _warn_and_use("quantit��", suggestions.get("quantita"), column_warnings)
    elif quantity_index is not None and not _column_has_values(data_rows, [quantity_index]):
        raise ValueError(
            f"La colonna quantit�� selezionata ({get_column_letter(quantity_index + 1)}) non contiene valori. "
            "Verifica la colonna scelta nel foglio."
        )
    if progressive_index is None:
        progressive_index = _warn_and_use("progressivo", suggestions.get("progressivo"), column_warnings)
    elif progressive_index is not None and not _column_has_values(data_rows, [progressive_index]):
        raise ValueError(
            f"La colonna progressivo selezionata ({get_column_letter(progressive_index + 1)}) non contiene valori. "
            "Verifica la colonna scelta nel foglio."
        )
    if not code_indexes and not description_indexes and progressive_index is None:
        raise ValueError(
            "Seleziona almeno una colonna da utilizzare come codice, descrizione o progressivo"
        )
    workbook_formulas = load_workbook(filename=file_path, data_only=False, read_only=True)
    try:
        formula_sheet = _select_sheet(workbook_formulas, sheet_name)
        raw_formula_rows = list(formula_sheet.iter_rows(min_row=header_idx + 2, values_only=False))
        formula_rows = _apply_column_filter(raw_formula_rows, kept_column_indexes)
    finally:
        workbook_formulas.close()

    formula_rows_iter = iter(formula_rows)

    voci: list[ParsedVoce] = []
    ordine = 0
    last_code: str | None = None
    last_desc: str | None = None
    last_progressivo: int | None = None
    for row in data_rows:
        formula_row = next(formula_rows_iter, ())
        if not _row_has_values(row):
            continue
        codice = _combine_code(row, code_indexes)
        descrizione = _combine_text(row, description_indexes)
        raw_price = _cell_to_float(row, price_index)
        quantita = _cell_to_float(row, quantity_index) if quantity_index is not None else None
        progressivo_value = _cell_to_progressive(row, progressive_index)

        is_totale_row = descrizione and "totale" in descrizione.lower()
        has_price = raw_price is not None

        # Se la riga contiene solo header (codice/descrizione/progressivo) senza quantità né prezzo,
        # memorizza il contesto e passa alla riga successiva.
        if not has_price and quantita is None:
            if codice or descrizione or progressivo_value is not None:
                last_code = codice or last_code
                last_desc = descrizione or last_desc
                if progressivo_value is not None:
                    last_progressivo = progressivo_value
            continue

        # Aggiorna il contesto corrente: le righe di descrizione/codice senza prezzo
        # fungono da header per la successiva riga "Totale".
        if (codice or descrizione) and not has_price and not quantita:
            last_code = codice or last_code
            last_desc = descrizione or last_desc
            if progressivo_value is not None:
                last_progressivo = progressivo_value

        if (codice or progressivo_value) and not is_totale_row and not has_price:
            last_code = codice or last_code
            last_desc = descrizione or last_desc
            if progressivo_value is not None:
                last_progressivo = progressivo_value
            continue

        if not is_totale_row and not codice and progressivo_value is None and not has_price:
            if descrizione:
                last_desc = descrizione
            continue

        # SKIP righe intermedie con solo quantità (senza prezzo, senza "Totale")
        # Queste righe sono dettagli parziali che precedono la riga "Totale" finale
        if not is_totale_row and not has_price and quantita is not None:
            # Mantieni il contesto ma non creare una voce
            continue

        # Gestione riga "Totale": usa codice/descrizione dell'header precedente,
        # ma prendi quantita/prezzo dalla riga totale.
        if is_totale_row and (quantita is not None or raw_price is not None):
            source_code = codice or last_code
            source_desc = last_desc or codice or "Voce senza descrizione"
            source_progressivo = progressivo_value or last_progressivo
            if source_progressivo is None:
                continue
            prezzo_value = _sanitize_price_candidate(raw_price) if raw_price is not None else None
            quantita_value = quantita if quantita not in (None,) else None
            importo_value = None
            if prezzo_value is not None and quantita_value is not None:
                _, importo_value = _calculate_line_amount(quantita_value, round(prezzo_value, 4))
                prezzo_value = round(prezzo_value, 4)
            elif raw_price is not None:
                importo_value = _ceil_amount(raw_price)

            wbs_levels: list[ParsedWbsLevel] = []
            normalized_code_value = _normalize_wbs7_code(source_code)
            if _looks_like_wbs7_code(normalized_code_value):
                wbs_levels.append(
                    ParsedWbsLevel(
                        level=7,
                        code=normalized_code_value,
                        description=source_desc,
                    )
                )
            voci.append(
                ParsedVoce(
                    ordine=ordine,
                    progressivo=source_progressivo,
                    codice=source_code,
                    descrizione=source_desc,
                    wbs_levels=wbs_levels,
                    unita_misura=None,
                    quantita=quantita_value,
                    prezzo_unitario=prezzo_value,
                    importo=importo_value,
                    note=None,
                    metadata=None,
                )
            )
            ordine += 1
            last_code = None
            last_desc = None
            last_progressivo = None
            # Non azzero last_code/last_desc per consentire utilizzo in mancanza di header successivo
            continue

        if not codice and last_code:
            codice = last_code
        if not descrizione and last_desc:
            descrizione = last_desc
        if raw_price is None:
            raw_price = 0.0
        if not codice and not descrizione and progressivo_value is None:
            continue
        if raw_price is None and quantita is None:
            continue
        if not codice and progressivo_value is not None:
            codice = f"PROG-{progressivo_value:05d}"
        if not descrizione and progressivo_value is not None:
            descrizione = f"Voce progressivo {progressivo_value}"
        last_code = None
        last_desc = None
        formula_cell = (
            formula_row[price_index] if price_index < len(formula_row) else None
        )
        if _has_external_formula(formula_cell):
            raise ValueError(
                "La colonna prezzo contiene formule collegate a file esterni. Apri il file in Excel e incolla i valori numerici prima di importare."
            )
        if quantity_index is not None and quantity_index < len(formula_row):
            quantity_formula_cell = formula_row[quantity_index]
        else:
            quantity_formula_cell = None
        if quantity_index is not None and _has_external_formula(quantity_formula_cell):
            raise ValueError(
                "La colonna quantit�� contiene formule collegate a file esterni. Incolla i valori numerici prima dell'import."
            )
        prezzo = _sanitize_price_candidate(raw_price)
        if prezzo is None:
            continue
        if is_totale_row:
            quantita_value = quantita if quantita not in (None,) else None
            prezzo_value: float | None = None
            importo_value = _ceil_amount(prezzo)
        else:
            quantita_value = quantita if quantita not in (None,) else None
            prezzo_value = round(prezzo, 4)
            _, importo_value = _calculate_line_amount(quantita_value, prezzo_value)
        voce_descrizione = descrizione or codice or "Voce senza descrizione"
        effective_progressivo = progressivo_value or last_progressivo
        if effective_progressivo is None:
            continue
        wbs_levels: list[ParsedWbsLevel] = []
        normalized_code_value = _normalize_wbs7_code(codice)
        if _looks_like_wbs7_code(normalized_code_value):
            wbs_levels.append(
                ParsedWbsLevel(
                    level=7,
                    code=normalized_code_value,
                    description=voce_descrizione,
                )
            )
        voci.append(
            ParsedVoce(
                ordine=ordine,
                progressivo=effective_progressivo,
                codice=codice,
                descrizione=voce_descrizione,
                wbs_levels=wbs_levels,
                unita_misura=None,
                quantita=quantita_value,
                prezzo_unitario=prezzo_value,
                importo=importo_value,
                note=None,
                metadata=None,
            )
        )
        ordine += 1

    computo = ParsedComputo(
        titolo=None,
        voci=voci,
        totale_quantita=None,
        totale_importo=None,
    )
    return _CustomReturnParseResult(computo=computo, column_warnings=column_warnings)


def _select_sheet(workbook, requested_name: str | None):
    if requested_name:
        if requested_name in workbook.sheetnames:
            return workbook[requested_name]
        raise ValueError(f"Il foglio {requested_name} non esiste nel file selezionato.")
    return workbook[workbook.sheetnames[0]]


def _rows_to_dataframe(rows: Sequence[Sequence[Any]]) -> pd.DataFrame:
    return pd.DataFrame(rows)


def _drop_empty_columns(rows: Sequence[Sequence[Any]]) -> tuple[list[list[Any]], int, list[int]]:
    """
    Rimuove colonne totalmente vuote usando pandas per evitare offset errati quando il file
    contiene colonne segnaposto o intere colonne vuote.
    Restituisce righe ripulite, numero colonne eliminate e gli indici originali mantenuti.
    """
    df = pd.DataFrame(rows)
    if df.empty:
        return [], 0, []

    # Keep columns that have at least one non-null/non-empty cell
    non_empty_mask = ~df.isna().all(axis=0)
    kept_columns = [int(col) for col, keep in zip(df.columns, non_empty_mask) if keep]
    cleaned = df.loc[:, non_empty_mask]

    cleaned_rows = cleaned.where(pd.notna(cleaned), None).values.tolist()
    dropped_count = len(df.columns) - len(kept_columns)
    return cleaned_rows, dropped_count, kept_columns


def _apply_column_filter(rows: Sequence[Sequence[Any]], kept_indexes: Sequence[int]) -> list[list[Any]]:
    if not kept_indexes:
        return [list(r) for r in rows]
    kept_set = set(kept_indexes)
    filtered: list[list[Any]] = []
    for row in rows:
        filtered.append([cell for idx, cell in enumerate(row) if idx in kept_set])
    return filtered


def _ratio(values: Sequence[str], predicate) -> float:
    if not values:
        return 0.0
    count = sum(1 for value in values if predicate(value))
    return count / len(values)


def _looks_numeric(value: str) -> bool:
    try:
        float(value)
        return True
    except ValueError:
        return False


def _looks_currency(value: str) -> bool:
    if not value:
        return False
    return bool(re.match(r"^\s*[-+]?[\d.,]+\s*$", value))


def _looks_code(value: str) -> bool:
    if not value:
        return False
    if len(value) < 3 or len(value) > 20:
        return False
    return bool(re.match(r"^[A-Za-z0-9._/-]+$", value))


def _looks_text(value: str) -> bool:
    if not value:
        return False
    if len(value) < 4:
        return False
    return bool(re.search(r"[A-Za-z]", value))


def _pick_column_profile(index: int, sample_rows: list[list[str]]) -> _ColumnProfile:
    values = [row[index] for row in sample_rows if len(row) > index]
    # Filtra valori vuoti (None, "", whitespace)
    clean_values = [
        value for value in values
        if value is not None and str(value).strip()
    ]
    text_values = [str(value) for value in clean_values if isinstance(value, str)]
    samples = text_values[:5]

    # Calcola empty_ratio: percentuale di celle vuote
    empty_ratio = 1.0 - (len(clean_values) / len(values)) if values else 1.0

    numeric_ratio = _ratio(text_values, _looks_numeric)
    currency_ratio = _ratio(text_values, _looks_currency)
    code_ratio = _ratio(text_values, _looks_code)
    text_ratio = _ratio(text_values, _looks_text)
    header_label = _normalize_header_text(values[0]) if values else None

    # Penalizza colonne molto vuote (> 70% vuote)
    empty_penalty = max(0, empty_ratio - 0.7) * 2  # da 0 a 0.6
    numeric_ratio = max(0, numeric_ratio - empty_penalty)
    currency_ratio = max(0, currency_ratio - empty_penalty)
    code_ratio = max(0, code_ratio - empty_penalty)
    text_ratio = max(0, text_ratio - empty_penalty)

    return _ColumnProfile(
        index=index,
        letter=get_column_letter(index + 1),
        header_label=header_label,
        samples=samples,
        numeric_ratio=numeric_ratio,
        currency_ratio=currency_ratio,
        code_ratio=code_ratio,
        text_ratio=text_ratio,
    )


def _detect_column_suggestions(rows: list, header_idx: int) -> dict[str, Any]:
    if not rows or header_idx < 0 or header_idx >= len(rows):
        return {}

    header_row = rows[header_idx]
    sample_rows = [
        [str(cell) if cell is not None else "" for cell in row]
        for row in rows[header_idx + 1 : header_idx + 11]
    ]
    profiles: list[_ColumnProfile] = []
    for idx in range(len(header_row)):
        profiles.append(_pick_column_profile(idx, sample_rows))

    suggestions: dict[str, _ColumnSuggestion] = {}

    def update_suggestion(target: str, candidate: _ColumnProfile, score: float):
        existing = suggestions.get(target)
        if existing is None or score > existing.score:
            suggestions[target] = _ColumnSuggestion(
                target=target,
                column_index=candidate.index,
                column_letter=candidate.letter,
                header_label=candidate.header_label,
                score=score,
            )

    for profile in profiles:
        header = (profile.header_label or "").lower()
        if "prezzo" in header or "importo" in header or profile.currency_ratio > 0.3:
            update_suggestion("prezzo", profile, profile.currency_ratio + 0.1)
        if "quant" in header or "q.t" in header or profile.numeric_ratio > 0.3:
            update_suggestion("quantita", profile, profile.numeric_ratio + 0.05)
        if "prog" in header:
            update_suggestion("progressivo", profile, profile.numeric_ratio + 0.05)
        if "cod" in header or profile.code_ratio > 0.3:
            update_suggestion("codice", profile, profile.code_ratio + 0.05)
        if "descr" in header or profile.text_ratio > 0.5:
            update_suggestion("descrizione", profile, profile.text_ratio + 0.05)

    return {"profiles": profiles, "suggestions": suggestions}


def _locate_header_row(rows: list) -> int | None:
    best_idx: int | None = None
    best_count = 0
    for idx, row in enumerate(rows[:30]):
        count = sum(1 for cell in row if _cell_has_content(cell))
        if count > best_count:
            best_idx = idx
            best_count = count
    return best_idx


def _columns_to_indexes(
    columns: Sequence[str],
    name: str,
    *,
    header_row: Sequence[str],
    required: bool = True,
) -> list[int]:
    if not columns:
        if required:
            raise ValueError(f"Seleziona almeno una colonna per {name}")
        return []

    indexes: list[int] = []
    for col in columns:
        normalized = _normalize_header_text(col)
        if normalized in (None,):
            continue
        try:
            indexes.append(column_index_from_string(normalized) - 1)
            continue
        except (TypeError, ValueError):
            pass
        for idx, header in enumerate(header_row):
            if _normalize_header_text(header) == normalized:
                indexes.append(idx)
                break
    if not indexes and required:
        raise ValueError(f"Nessuna colonna valida fornita per {name}")
    return indexes


def _single_column_index(
    column: str | None,
    name: str,
    *,
    header_row: Sequence[str],
    required: bool = True,
) -> int | None:
    if column is None:
        if required:
            raise ValueError(f"Seleziona una colonna per {name}")
        return None
    normalized = _normalize_header_text(column)
    try:
        return column_index_from_string(normalized) - 1
    except (TypeError, ValueError):
        pass
    for idx, header in enumerate(header_row):
        if _normalize_header_text(header) == normalized:
            return idx
    if required:
        raise ValueError(f"Colonna {name} non trovata: {column}")
    return None


def _resolve_column_reference(reference: str, header_row: Sequence[str]) -> int:
    normalized = _normalize_header_text(reference)
    try:
        return column_index_from_string(normalized) - 1
    except (TypeError, ValueError):
        pass
    for idx, header in enumerate(header_row):
        if _normalize_header_text(header) == normalized:
            return idx
    raise ValueError(f"Colonna non trovata: {reference}")


def _normalize_header_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip().lower()


def _row_has_values(row) -> bool:
    return any(_cell_has_content(cell) for cell in row)


def _column_has_values(rows: Sequence[Sequence[Any]], indexes: Sequence[int]) -> bool:
    for row in rows:
        for idx in indexes:
            if idx < len(row) and _cell_has_content(row[idx]):
                return True
    return False


def _sanitize_price_candidate(value: float) -> float | None:
    if value is None:
        return None
    try:
        decimal_value = Decimal(str(value))
        return float(decimal_value)
    except (ValueError, TypeError, ArithmeticError):
        return None


def _has_external_formula(cell) -> bool:
    if cell is None:
        return False
    if not hasattr(cell, "value"):
        return False
    value = cell.value
    if value is None or not isinstance(value, str):
        return False
    if "!" in value and "[" in value:
        return True
    return False


def _cell_has_content(value) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        return bool(value.strip())
    return True


def _combine_text(row, indexes: Sequence[int]) -> str | None:
    values: list[str] = []
    for idx in indexes:
        if idx < len(row):
            text = _cell_to_text(row[idx])
            if text:
                values.append(text)
    combined = " ".join(values).strip()
    return combined or None


def _combine_code(row, indexes: Sequence[int]) -> str | None:
    for idx in indexes:
        if idx < len(row):
            text = _cell_to_text(row[idx])
            code = _extract_code_from_text(text)
            if code:
                return code
    return None


def _cell_to_text(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, (int, float, Decimal)):
        return str(value)
    text = str(value).strip()
    return text or None


def _extract_code_from_text(text: str | None) -> str | None:
    if not text:
        return None
    candidates = re.findall(r"[A-Za-z0-9][A-Za-z0-9._/-]{1,19}", text)
    if candidates:
        return candidates[0]
    return None


def _cell_to_float(row, index: int) -> float | None:
    if index is None or index < 0 or index >= len(row):
        return None
    value = row[index]
    if value in (None, "", " "):
        return None
    if isinstance(value, (int, float, Decimal)):
        try:
            return float(value)
        except (TypeError, ValueError):
            return None
    text = str(value).strip()
    if not text:
        return None
    text = text.replace(",", ".")
    text = text.replace(" ", "")
    try:
        return float(text)
    except ValueError:
        return None


def _cell_to_progressive(row, index: int | None) -> int | None:
    if index is None or index < 0 or index >= len(row):
        return None
    value = row[index]
    if value in (None, "", " "):
        return None
    if isinstance(value, (int, float, Decimal)):
        try:
            return int(float(value))
        except (TypeError, ValueError):
            return None
    text = str(value).strip()
    if not text:
        return None
    text = text.replace(",", ".")
    try:
        return int(float(text))
    except ValueError:
        return None


@dataclass
class _ColumnProfile:
    index: int
    letter: str
    header_label: str | None
    samples: list[str]
    numeric_ratio: float
    currency_ratio: float
    code_ratio: float
    text_ratio: float


@dataclass
class _ColumnSuggestion:
    target: str
    column_index: int
    column_letter: str
    header_label: str | None
    score: float


@dataclass
class _CustomReturnParseResult:
    computo: ParsedComputo
    column_warnings: list[str]
