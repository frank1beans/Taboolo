from __future__ import annotations

import logging
import unicodedata
from collections import defaultdict, deque
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP, ROUND_UP
import copy
import re
from pathlib import Path
from typing import Any, Iterable, Sequence, Dict, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from sqlmodel import Session, select

from app.db.models import Computo, ComputoTipo, PriceListItem, PriceListOffer, VoceComputo
from app.db.models_wbs import (
    Impresa,
    Voce as VoceNorm,
    VoceOfferta,
    VoceProgetto,
    Wbs6,
    Wbs7,
    WbsSpaziale,
)
from app.excel import ParsedComputo, ParsedVoce, ParsedWbsLevel, parse_computo_excel
from app.excel.parser import MAX_WBS_LEVELS
from app.services.commesse import CommesseService
from app.services.nlp import semantic_embedding_service

logger = logging.getLogger(__name__)


@dataclass
class _ReturnAlignmentResult:
    voci_allineate: list[ParsedVoce]
    legacy_pairs: list[tuple[VoceComputo | None, ParsedVoce]]
    matched_count: int
    price_adjustments: list[str]
    zero_guard_inputs: list[ParsedVoce]
    return_only_labels: list[str]
    progress_quantity_mismatches: list[str]
    progress_price_conflicts: list[str]
    excel_only_groups: list[str]


@dataclass
class _ColumnProfile:
    index: int
    letter: str
    header_label: str | None
    samples: list[str]
    numeric_ratio: float
    currency_ratio: float
    code_ratio: float
    text_ratio: float


@dataclass
class _ColumnSuggestion:
    target: str
    column_index: int
    column_letter: str
    header_label: str | None
    score: float


@dataclass
class _CustomReturnParseResult:
    computo: ParsedComputo
    column_warnings: list[str]


_CODE_REGEX = re.compile(
    r"^[A-Z]{1,4}\d{1,4}(?:[.\-_/]?[A-Z0-9]{1,4})*$",
    re.IGNORECASE,
)


class ImportService:
    """Gestisce il caricamento di file Excel e la creazione delle voci di computo."""

    @staticmethod
    def _sanitize_impresa_label(label: str | None) -> str | None:
        """Normalizza il nome impresa rimuovendo suffissi duplicati e spazi superflui."""
        if not label:
            return None
        text = label.strip()
        if not text:
            return None
        text = re.sub(r"\s*\(\d+\)\s*$", "", text).strip()
        return text or None

    @staticmethod
    def _get_or_create_impresa(session: Session, label: str | None) -> Optional[Impresa]:
        """Recupera o crea un'impresa normalizzando il nome."""
        if not label:
            return None
        text = ImportService._sanitize_impresa_label(label)
        if not text:
            return None
        normalized = re.sub(r"\s+", " ", text).lower()
        if not normalized:
            return None
        existing = session.exec(
            select(Impresa).where(Impresa.normalized_label == normalized)
        ).first()
        if existing:
            return existing
        impresa = Impresa(label=text, normalized_label=normalized)
        session.add(impresa)
        session.flush()
        return impresa

    def import_computo_progetto(
        self,
        *,
        session: Session,
        commessa_id: int,
        file: Path,
        originale_nome: str | None,
    ) -> Computo:
        commessa = CommesseService.get_commessa(session, commessa_id)
        if not commessa:
            raise ValueError("Commessa non trovata")

        parser_result = parse_computo_excel(file)
        if parser_result.totale_importo is not None:
            total_import = _ceil_amount(parser_result.totale_importo)
        else:
            total_import = _ceil_amount(sum(voce.importo or 0 for voce in parser_result.voci))

        computo = CommesseService.add_computo(
            session,
            commessa,
            nome=f"{commessa.nome} - Computo progetto",
            tipo=ComputoTipo.progetto,
            file_nome=originale_nome,
            file_percorso=str(file),
        )

        computo.importo_totale = total_import
        session.add(computo)
        self.persist_project_from_parsed(
            session=session,
            commessa_id=commessa.id,
            computo=computo,
            parsed_voci=parser_result.voci,
        )
        session.commit()
        session.refresh(computo)
        return computo

    def _import_lc_return(
        self,
        *,
        session: Session,
        commessa,
        parser_result: ParsedComputo,
        column_warnings: Sequence[str] | None = None,
        progetto_voci: Sequence[VoceComputo],
        impresa: str,
        impresa_id: int | None,
        file: Path,
        originale_nome: str | None,
        resolved_round: int,
        target_computo: Computo | None,
    ) -> Computo:
        raw_entries = parser_result.voci
        if not raw_entries:
            raise ValueError("Il file LC non contiene voci utilizzabili.")

        if target_computo is not None:
            computo = target_computo
            computo.file_nome = originale_nome
            computo.file_percorso = str(file)
            computo.impresa = impresa
            computo.impresa_id = impresa_id
        else:
            computo = CommesseService.add_computo(
                session,
                commessa,
                nome=f"{commessa.nome} - {impresa}",
                tipo=ComputoTipo.ritorno,
                impresa=impresa,
                impresa_id=impresa_id,
                round_number=resolved_round,
                file_nome=originale_nome,
                file_percorso=str(file),
            )

        computo.round_number = resolved_round
        computo.file_nome = originale_nome
        computo.file_percorso = str(file)
        computo.updated_at = datetime.utcnow()
        session.add(computo)

        summary = self._sync_price_list_offers(
            session=session,
            commessa_id=commessa.id,
            computo=computo,
            impresa_label=impresa,
            parsed_voci=raw_entries,
            collect_summary=True,
        )
        if not summary or summary["price_items_total"] == 0:
            raise ValueError(
                "Nessun elenco prezzi associato alla commessa: impossibile importare il file LC."
            )

        matching_report = _build_lc_matching_report(summary)
        computo.matching_report = matching_report

        missing_count = max(
            0,
            summary["price_items_total"] - len(summary["matched_item_ids"]),
        )
        note_messages: list[str] = []
        if missing_count:
            note_messages.append(
                f"{missing_count} voci del listino non hanno trovato un prezzo nel file LC."
            )
        if column_warnings:
            note_messages.extend(column_warnings)
        computo.note = " ".join(note_messages) if note_messages else None

        price_items = summary.get("price_items") or []
        price_map: dict[int, float] = summary.get("price_map") or {}
        voci_allineate, legacy_pairs = _build_project_snapshot_from_price_offers(
            progetto_voci,
            price_items,
            price_map,
        )

        total_import = _ceil_amount(
            sum(voce.importo or 0 for voce in voci_allineate)
        )
        computo.importo_totale = total_import or 0.0

        self._bulk_insert_voci(session, computo, voci_allineate)
        if legacy_pairs:
            self._sync_normalized_offerte(
                session,
                commessa.id,
                computo,
                impresa,
                resolved_round,
                legacy_pairs,
            )

        session.commit()
        session.refresh(computo)
        return computo

    def import_computo_ritorno(
        self,
        *,
        session: Session,
        commessa_id: int,
        impresa: str,
        file: Path,
        originale_nome: str | None,
        round_number: int | None = None,
        round_mode: str = "auto",
        sheet_name: str | None = None,
        sheet_code_columns: Sequence[str] | None = None,
        sheet_description_columns: Sequence[str] | None = None,
        sheet_price_column: str | None = None,
        sheet_quantity_column: str | None = None,
        sheet_progressive_column: str | None = None,
    ) -> Computo:
        commessa = CommesseService.get_commessa(session, commessa_id)
        if not commessa:
            raise ValueError("Commessa non trovata")

        progetto = session.exec(
            select(Computo)
            .where(
                Computo.commessa_id == commessa_id,
                Computo.tipo == ComputoTipo.progetto,
            )
            .order_by(Computo.created_at.desc())
        ).first()
        if not progetto:
            raise ValueError(
                "Carica prima un computo metrico estimativo per la commessa"
            )

        impresa = self._sanitize_impresa_label(impresa)
        impresa_entry = self._get_or_create_impresa(session, impresa)

        progetto_voci = session.exec(
            select(VoceComputo)
            .where(VoceComputo.computo_id == progetto.id)
            .order_by(VoceComputo.ordine)
        ).all()
        if not progetto_voci:
            raise ValueError(
                "Il computo metrico estimativo non contiene voci importabili"
            )

        existing_ritorni = session.exec(
            select(Computo)
            .where(
                Computo.commessa_id == commessa_id,
                Computo.tipo == ComputoTipo.ritorno,
                Computo.impresa == impresa,
            )
            .order_by(Computo.created_at.asc())
        ).all()

        normalized_mode = (round_mode or "auto").strip().lower()
        if normalized_mode not in {"auto", "new", "replace"}:
            raise ValueError("Modalità round non valida. Usa 'new' oppure 'replace'.")

        target_computo: Computo | None = None
        resolved_round: int

        if normalized_mode == "replace":
            if round_number is None:
                raise ValueError("Seleziona il round da aggiornare.")
            target_computo = next(
                (item for item in existing_ritorni if item.round_number == round_number),
                None,
            )
            if target_computo is None:
                raise ValueError(
                    f"Nessun computo dell'impresa {impresa} trovato per il round {round_number}."
                )
            resolved_round = round_number
        else:
            if round_number is not None:
                resolved_round = round_number
            else:
                existing_numbers = [
                    item.round_number
                    for item in existing_ritorni
                    if item.round_number is not None
                ]
                if existing_numbers:
                    resolved_round = max(existing_numbers) + 1
                else:
                    resolved_round = 1

            if any(
                item.round_number == resolved_round for item in existing_ritorni
            ):
                raise ValueError(
                    f"Esiste già un computo dell'impresa {impresa} per il round {resolved_round}. "
                    "Scegli la modalità di aggiornamento oppure seleziona un round diverso."
                )

        lc_mode = bool(sheet_price_column)
        if lc_mode:
            parse_result = _parse_custom_return_excel(
                file,
                sheet_name,
                sheet_code_columns or [],
                sheet_description_columns or [],
                sheet_price_column or "",
                sheet_quantity_column,
                sheet_progressive_column,
            )
            parser_result = parse_result.computo
            return self._import_lc_return(
                session=session,
                commessa=commessa,
                parser_result=parser_result,
                column_warnings=parse_result.column_warnings,
                progetto_voci=progetto_voci,
                impresa=impresa,
                impresa_id=impresa_entry.id if impresa_entry else None,
                file=file,
                originale_nome=originale_nome,
                resolved_round=resolved_round,
                target_computo=target_computo,
            )

        parser_result = parse_computo_excel(file, sheet_name=sheet_name)
        description_price_map = _build_description_price_map(parser_result.voci)
        ritorno_con_progressivi = _has_progressivi(parser_result.voci)
        progetto_quantita_totale = _sum_project_quantities(progetto_voci)
        excel_quantita_totale = (
            Decimal(str(parser_result.totale_quantita))
            if parser_result.totale_quantita is not None
            else None
        )
        ritorno_quantita_totale = (
            excel_quantita_totale
            if excel_quantita_totale is not None
            else sum(Decimal(str(voce.quantita or 0)) for voce in parser_result.voci)
        )
        progetto_quantita_float = float(progetto_quantita_totale or 0)
        ritorno_quantita_float = float(ritorno_quantita_totale or 0)
        delta_quantita_totale = ritorno_quantita_float - progetto_quantita_float

        total_voci = len(progetto_voci)
        alignment = _align_return_rows(
            progetto_voci,
            parser_result.voci,
            prefer_progressivi=ritorno_con_progressivi,
            description_price_map=description_price_map,
        )

        voci_allineate = alignment.voci_allineate
        legacy_pairs = alignment.legacy_pairs
        matched_count = alignment.matched_count
        price_adjustments = alignment.price_adjustments
        zero_guard_inputs = alignment.zero_guard_inputs
        return_only_labels = alignment.return_only_labels
        progress_quantity_mismatches = alignment.progress_quantity_mismatches
        progress_price_conflicts = alignment.progress_price_conflicts
        excel_only_groups = alignment.excel_only_groups
        matching_report = _build_matching_report(
            legacy_pairs=legacy_pairs,
            excel_only_labels=return_only_labels,
            excel_only_groups=excel_only_groups,
            quantity_mismatches=progress_quantity_mismatches if ritorno_con_progressivi else None,
            quantity_totals={
                "progetto": progetto_quantita_float,
                "ritorno": ritorno_quantita_float,
                "delta": delta_quantita_totale,
            },
        )

        warning_message: str | None = None
        remaining_missing = [
            _voce_label(voce_progetto)
            for voce_progetto, parsed in legacy_pairs
            if voce_progetto
            and (parsed.metadata or {}).get("missing_from_return")
        ]
        if ritorno_con_progressivi and progress_quantity_mismatches:
            summary = "; ".join(progress_quantity_mismatches[:5])
            quantity_note = (
                f"{len(progress_quantity_mismatches)} progressivi riportano quantità diverse "
                f"rispetto al computo: {summary}"
            )
            if warning_message:
                warning_message += " " + quantity_note
            else:
                warning_message = quantity_note
        if ritorno_con_progressivi and progress_price_conflicts:
            summary = "; ".join(progress_price_conflicts[:5])
            price_note = (
                f"{len(progress_price_conflicts)} progressivi hanno prezzi non coerenti: {summary}"
            )
            if warning_message:
                warning_message += " " + price_note
            else:
                warning_message = price_note
        if remaining_missing:
            elenco = ", ".join(
                _shorten_label(item) for item in remaining_missing[:5]
            )
            if len(remaining_missing) > 5:
                elenco += ", ..."
            if matched_count == 0:
                raise ValueError(
                    "Nessuna voce del file caricato corrisponde al computo metrico estimativo. "
                    "Verifica di avere esportato il ritorno con la stessa struttura del computo di progetto: "
                    f"{elenco}"
                )
            if total_voci:
                coverage = matched_count / total_voci
                warning_message = (
                    f"{len(remaining_missing)} voci del computo metrico non sono state aggiornate dal ritorno "
                    f"(allineate {matched_count} su {total_voci}, prime: {elenco})"
                )
                if coverage < 0.5:
                    warning_message += ". Il file sembra fornire solo una parte delle voci."
            else:
                warning_message = (
                    f"{len(remaining_missing)} voci del computo metrico non sono state aggiornate dal ritorno: {elenco}"
                )

        if excel_only_groups:
            elenco_excel = ", ".join(
                _shorten_label(item) for item in excel_only_groups[:5]
            )
            if len(excel_only_groups) > 5:
                elenco_excel += ", ..."
            extra_warning = (
                f"{len(excel_only_groups)} voci del ritorno non sono state abbinate al computo: {elenco_excel}"
            )
            if warning_message:
                warning_message += " " + extra_warning
            else:
                warning_message = extra_warning
        if price_adjustments:
            adjustments_summary = "; ".join(price_adjustments[:5])
            if len(price_adjustments) > 5:
                adjustments_summary += ", ..."
            adjustments_text = (
                "Corrette automaticamente alcune offerte con prezzi fuori scala: "
                f"{adjustments_summary}"
            )
            if warning_message:
                warning_message += " " + adjustments_text
            else:
                warning_message = adjustments_text

        if return_only_labels:
            extras_summary = ", ".join(return_only_labels[:5])
            if len(return_only_labels) > 5:
                extras_summary += ", ..."
            extra_warning = (
                f"Importate {len(return_only_labels)} voci presenti solo nel ritorno di gara: {extras_summary}"
            )
            if warning_message:
                warning_message += " " + extra_warning
            else:
                warning_message = extra_warning

        zero_guard_violations = _detect_forced_zero_violations(zero_guard_inputs)
        if zero_guard_violations:
            summary = ", ".join(zero_guard_violations[:5])
            if len(zero_guard_violations) > 5:
                summary += ", ..."
            zero_guard_warning = (
                "Alcune voci di coordinamento (Assistenze murarie / Mark up fee) risultano valorizzate "
                f"ma devono restare a zero: {summary}. Correggi il file del ritorno."
            )
            if warning_message:
                warning_message += " " + zero_guard_warning
            else:
                warning_message = zero_guard_warning

        total_import: float | None = None
        computed_total: Decimal | None = None
        if voci_allineate:
            computed_total = sum(
                Decimal(str(voce.importo))
                for voce in voci_allineate
                if voce.importo is not None
            )
            computed_total = computed_total.quantize(
                Decimal("0.01"), rounding=ROUND_HALF_UP
            )
            total_import = float(computed_total)

        if parser_result.totale_importo is not None:
            excel_total = Decimal(str(parser_result.totale_importo)).quantize(
                Decimal("0.01"), rounding=ROUND_HALF_UP
            )
            if computed_total is None or abs(excel_total - computed_total) <= Decimal("0.01"):
                total_import = float(excel_total)
            else:
                extra_warning = (
                    "Il totale del ritorno "
                    f"({format(excel_total, '.2f')}) non coincide con la somma delle voci importate "
                    f"({format(computed_total, '.2f')})."
                )
                if warning_message:
                    warning_message += " " + extra_warning
                else:
                    warning_message = extra_warning

        if excel_quantita_totale is not None and progetto_quantita_totale is not None:
            excel_quantity = excel_quantita_totale.quantize(
                Decimal("0.0001"), rounding=ROUND_HALF_UP
            )
            progetto_quantity = progetto_quantita_totale.quantize(
                Decimal("0.0001"), rounding=ROUND_HALF_UP
            )
            if abs(excel_quantity - progetto_quantity) > Decimal("0.0001"):
                quantity_warning = (
                    "Il totale delle quantità importate "
                    f"({_format_quantity_value(excel_quantity)}) non coincide con il computo metrico "
                    f"({_format_quantity_value(progetto_quantity)})."
                )
                if warning_message:
                    warning_message += " " + quantity_warning
                else:
                    warning_message = quantity_warning

        if target_computo is not None:
            computo = target_computo
            computo.file_nome = originale_nome
            computo.file_percorso = str(file)
            computo.impresa = impresa
        else:
            computo = CommesseService.add_computo(
                session,
                commessa,
                nome=f"{commessa.nome} - {impresa}",
                tipo=ComputoTipo.ritorno,
                impresa=impresa,
                round_number=resolved_round,
                file_nome=originale_nome,
                file_percorso=str(file),
            )

        computo.round_number = resolved_round
        computo.importo_totale = total_import
        computo.file_nome = originale_nome
        computo.file_percorso = str(file)
        if warning_message:
            computo.note = warning_message
        else:
            computo.note = None
        computo.updated_at = datetime.utcnow()
        computo.matching_report = matching_report
        session.add(computo)
        self._bulk_insert_voci(session, computo, voci_allineate)
        self._sync_normalized_offerte(
            session,
            commessa.id,
            computo,
            impresa,
            resolved_round,
            legacy_pairs,
        )
        self._sync_normalized_offerte(
            session,
            commessa.id,
            computo,
            impresa,
            resolved_round,
            legacy_pairs,
        )
        self._sync_price_list_offers(
            session=session,
            commessa_id=commessa.id,
            computo=computo,
            impresa_label=impresa,
            parsed_voci=voci_allineate,
        )
        session.commit()
        session.refresh(computo)
        return computo

    def update_manual_offer_price(
        self,
        *,
        session: Session,
        commessa_id: int,
        computo_id: int,
        price_list_item_id: int,
        prezzo_unitario: float,
        quantita: float | None = None,
    ) -> tuple[PriceListOffer, Computo]:
        computo = session.get(Computo, computo_id)
        if not computo or computo.commessa_id != commessa_id:
            raise ValueError("Computo non valido per la commessa selezionata.")
        if computo.tipo != ComputoTipo.ritorno:
            raise ValueError("L'aggiornamento manuale dei prezzi è consentito solo sui ritorni gara.")

        item = session.get(PriceListItem, price_list_item_id)
        if not item or item.commessa_id != commessa_id:
            raise ValueError("Voce di elenco prezzi non trovata per la commessa indicata.")

        context = _WbsNormalizeContext(session, commessa_id)
        impresa_entry = (
            context.get_or_create_impresa(computo.impresa) if computo.impresa else None
        )
        offer = session.exec(
            select(PriceListOffer).where(
                PriceListOffer.computo_id == computo.id,
                PriceListOffer.price_list_item_id == price_list_item_id,
            )
        ).first()
        if not offer:
            offer = PriceListOffer(
                price_list_item_id=price_list_item_id,
                commessa_id=commessa_id,
                computo_id=computo.id,
                impresa_id=impresa_entry.id if impresa_entry else None,
                impresa_label=computo.impresa,
                round_number=computo.round_number,
            )
        offer.prezzo_unitario = round(float(prezzo_unitario), 4)
        if quantita is not None:
            offer.quantita = quantita
        offer.updated_at = datetime.utcnow()
        session.add(offer)

        self._rebuild_computo_from_offers(session, commessa_id, computo)
        self._acknowledge_manual_price(computo, price_list_item_id)

        session.commit()
        session.refresh(offer)
        session.refresh(computo)
        return offer, computo

    def persist_project_from_parsed(
        self,
        *,
        session: Session,
        commessa_id: int,
        computo: Computo,
        parsed_voci: Sequence[ParsedVoce],
    ) -> list[VoceComputo]:
        """
        Salva le voci di un computo di progetto partendo da ParsedVoce già forniti (es. da file SIX).
        """
        legacy_voci = self._bulk_insert_voci(session, computo, parsed_voci)
        self._sync_normalized_progetto(
            session,
            commessa_id,
            computo,
            legacy_voci,
            parsed_voci,
        )
        return legacy_voci

    def _bulk_insert_voci(
        self,
        session: Session,
        computo: Computo,
        voci: Iterable[ParsedVoce],
    ) -> list[VoceComputo]:
        session.exec(
            VoceComputo.__table__.delete().where(VoceComputo.computo_id == computo.id)
        )

        voce_models = []
        commessa_id = computo.commessa_id
        commessa_code = getattr(computo, "commessa_code", None)
        commessa_tag = _normalize_commessa_tag(commessa_id, commessa_code)
        for parsed in voci:
            wbs_kwargs = _map_wbs_levels(parsed.wbs_levels)
            global_code = _build_global_voce_code(commessa_tag, parsed)

            voce_models.append(
                VoceComputo(
                    commessa_id=commessa_id,
                    commessa_code=commessa_code,
                    computo_id=computo.id,
                    global_code=global_code,
                    progressivo=parsed.progressivo,
                    codice=parsed.codice,
                    descrizione=parsed.descrizione,
                    unita_misura=parsed.unita_misura,
                    quantita=parsed.quantita,
                    prezzo_unitario=parsed.prezzo_unitario,
                    importo=parsed.importo,
                    note=parsed.note,
                    ordine=parsed.ordine,
                    extra_metadata=parsed.metadata,
                    **wbs_kwargs,
                )
            )

        session.add_all(voce_models)
        session.flush()
        return voce_models

    def _sync_normalized_progetto(
        self,
        session: Session,
        commessa_id: int,
        computo: Computo,
        legacy_voci: Sequence[VoceComputo],
        parsed_voci: Sequence[ParsedVoce],
    ) -> None:
        context = _WbsNormalizeContext(session, commessa_id)
        session.exec(
            VoceProgetto.__table__.delete().where(VoceProgetto.computo_id == computo.id)
        )
        for legacy, parsed in zip(legacy_voci, parsed_voci):
            voce_norm = context.ensure_voce(parsed, legacy)
            if voce_norm is None:
                continue
            session.add(
                VoceProgetto(
                    voce_id=voce_norm.id,
                    computo_id=computo.id,
                    quantita=parsed.quantita,
                    prezzo_unitario=parsed.prezzo_unitario,
                    importo=parsed.importo,
                    note=parsed.note,
                )
            )

    def _sync_normalized_offerte(
        self,
        session: Session,
        commessa_id: int,
        computo: Computo,
        impresa_label: str,
        round_number: int | None,
        legacy_pairs: Sequence[tuple[VoceComputo | None, ParsedVoce]],
    ) -> None:
        context = _WbsNormalizeContext(session, commessa_id)
        impresa_entry = context.get_or_create_impresa(impresa_label)
        if not impresa_entry:
            return
        session.exec(
            VoceOfferta.__table__.delete().where(VoceOfferta.computo_id == computo.id)
        )
        for legacy, parsed in legacy_pairs:
            voce_norm = context.get_voce_from_legacy(legacy.id) if legacy else None
            if voce_norm is None:
                voce_norm = context.ensure_voce(parsed, legacy)
            if voce_norm is None:
                continue
            session.add(
                VoceOfferta(
                    voce_id=voce_norm.id,
                    computo_id=computo.id,
                    impresa_id=impresa_entry.id,
                    round_number=round_number,
                    quantita=parsed.quantita,
                    prezzo_unitario=parsed.prezzo_unitario,
                    importo=parsed.importo,
                    note=parsed.note,
                )
            )

    def _sync_price_list_offers(
        self,
        session: Session,
        commessa_id: int,
        computo: Computo,
        impresa_label: str | None,
        parsed_voci: Sequence[ParsedVoce],
        *,
        persist_impresa: bool = True,
        collect_summary: bool = False,
    ) -> dict[str, Any] | None:
        session.exec(
            PriceListOffer.__table__.delete().where(
                PriceListOffer.computo_id == computo.id
            )
        )
        if not parsed_voci:
            return {
                "price_items_total": 0,
                "matched_item_ids": set(),
                "unmatched_entries": [],
                "price_items": [],
            } if collect_summary else None

        price_items = session.exec(
            select(PriceListItem).where(PriceListItem.commessa_id == commessa_id)
        ).all()
        if not price_items:
            return {
                "price_items_total": 0,
                "matched_item_ids": set(),
                "unmatched_entries": list(parsed_voci),
                "price_items": [],
            } if collect_summary else None

        (
            code_map,
            signature_map,
            description_map,
            head_signature_map,
            tail_signature_map,
            embedding_map,
        ) = _build_price_list_lookup(price_items)
        context = _WbsNormalizeContext(session, commessa_id)
        impresa_entry = (
            context.get_or_create_impresa(impresa_label)
            if persist_impresa and impresa_label
            else None
        )

        processed_ids: set[int] = set()
        matched_item_ids: set[int] = set()
        unmatched_entries: list[ParsedVoce] = []
        offer_models: list[PriceListOffer] = []
        price_map: dict[int, float] = {}
        price_records: dict[int, list[dict[str, Any]]] = defaultdict(list)
        conflicting_price_items: dict[int, dict[str, Any]] = {}

        for voce in parsed_voci:
            prezzo = voce.prezzo_unitario
            if prezzo in (None,):
                continue
            target_item = _match_price_list_item_entry(
                voce,
                code_map,
                signature_map,
                description_map,
                head_signature_map,
                tail_signature_map,
                embedding_map,
            )
            if not target_item:
                unmatched_entries.append(voce)
                continue

            price_value = round(float(prezzo), 4)
            source_label = voce.codice or voce.descrizione or "Voce senza descrizione"
            samples = price_records[target_item.id]
            samples.append({"source": source_label, "price": price_value})

            existing_price = price_map.get(target_item.id)
            if existing_price is not None and not _prices_match(existing_price, price_value):
                conflict_entry = conflicting_price_items.setdefault(
                    target_item.id,
                    {
                        "price_list_item_id": target_item.id,
                        "item_code": target_item.item_code,
                        "item_description": target_item.item_description,
                        "prices": set(),
                        "samples": list(samples),
                    },
                )
                conflict_entry["prices"].add(existing_price)
                conflict_entry["prices"].add(price_value)
                conflict_entry["samples"] = list(samples)
                logger.warning(
                    "Price conflict for code %s (item %s): existing=%s, new=%s",
                    target_item.item_code or target_item.product_id,
                    target_item.id,
                    existing_price,
                    price_value,
                )
                continue

            if target_item.id in processed_ids:
                continue

            processed_ids.add(target_item.id)
            matched_item_ids.add(target_item.id)
            price_map[target_item.id] = price_value
            offer_models.append(
                PriceListOffer(
                    price_list_item_id=target_item.id,
                    commessa_id=commessa_id,
                    computo_id=computo.id,
                    impresa_id=impresa_entry.id if impresa_entry else None,
                    impresa_label=impresa_label,
                    round_number=computo.round_number,
                    prezzo_unitario=round(float(prezzo), 4),
                    quantita=voce.quantita,
                )
            )

        if unmatched_entries:
            _log_unmatched_price_entries(unmatched_entries)
        if conflicting_price_items:
            _log_price_conflicts(conflicting_price_items.values())

        if offer_models:
            session.add_all(offer_models)
        if collect_summary:
            conflict_payload = [
                {
                    "price_list_item_id": entry["price_list_item_id"],
                    "item_code": entry["item_code"],
                    "item_description": entry["item_description"],
                    "prices": sorted({round(float(p), 4) for p in entry["prices"]}),
                    "samples": entry["samples"],
                }
                for entry in conflicting_price_items.values()
            ]
            return {
                "price_items_total": len(price_items),
                "matched_item_ids": matched_item_ids,
                "unmatched_entries": unmatched_entries,
                "price_items": price_items,
                "price_map": price_map,
                "conflicting_price_items": conflict_payload,
            }
        return None

    def _rebuild_computo_from_offers(
        self,
        session: Session,
        commessa_id: int,
        computo: Computo,
    ) -> None:
        progetto = session.exec(
            select(Computo)
            .where(
                Computo.commessa_id == commessa_id,
                Computo.tipo == ComputoTipo.progetto,
            )
            .order_by(Computo.created_at.desc())
        ).first()
        if not progetto:
            raise ValueError("Impossibile aggiornare il ritorno: manca il computo di progetto.")

        progetto_voci = session.exec(
            select(VoceComputo)
            .where(VoceComputo.computo_id == progetto.id)
            .order_by(VoceComputo.ordine)
        ).all()
        if not progetto_voci:
            raise ValueError("Il computo di progetto non contiene voci da utilizzare.")

        price_items = session.exec(
            select(PriceListItem).where(PriceListItem.commessa_id == commessa_id)
        ).all()
        offer_entries = session.exec(
            select(PriceListOffer).where(PriceListOffer.computo_id == computo.id)
        ).all()
        price_map = {
            entry.price_list_item_id: entry.prezzo_unitario
            for entry in offer_entries
            if entry.prezzo_unitario is not None
        }
        voci_allineate, legacy_pairs = _build_project_snapshot_from_price_offers(
            progetto_voci,
            price_items,
            price_map,
        )
        total_import = _ceil_amount(sum(voce.importo or 0 for voce in voci_allineate))
        computo.importo_totale = total_import or 0.0
        computo.updated_at = datetime.utcnow()
        self._bulk_insert_voci(session, computo, voci_allineate)
        if legacy_pairs:
            self._sync_normalized_offerte(
                session,
                commessa_id,
                computo,
                computo.impresa or "",
                computo.round_number,
                legacy_pairs,
            )

    @staticmethod
    def _acknowledge_manual_price(computo: Computo, price_list_item_id: int) -> None:
        report = computo.matching_report
        if not isinstance(report, dict):
            return
        if report.get("mode") != "lc":
            return
        missing_items = report.get("missing_price_items")
        if not isinstance(missing_items, list):
            return
        updated_missing = [
            item
            for item in missing_items
            if item.get("price_list_item_id") != price_list_item_id
        ]
        if len(updated_missing) == len(missing_items):
            return
        report["missing_price_items"] = updated_missing
        total = report.get("total_price_items")
        if isinstance(total, int) and total >= 0:
            report["matched_price_items"] = total - len(updated_missing)
        else:
            current_matched = report.get("matched_price_items") or 0
            report["matched_price_items"] = current_matched + 1
        computo.matching_report = report



def _collect_return_only_labels(
    wrappers: Sequence[dict[str, Any]],
    satisfied_group_keys: set[str] | None = None,
) -> list[str]:
    matched_groups = satisfied_group_keys or set()
    labels: list[str] = []
    for wrapper in wrappers:
        if wrapper.get("matched"):
            continue
        voce_excel = wrapper.get("voce")
        if voce_excel is None:
            continue
        key = _wbs_key_from_parsed(voce_excel)
        if key and key in matched_groups:
            continue
        label = voce_excel.descrizione or voce_excel.codice or "voce senza descrizione"
        labels.append(_shorten_label(label))
    return labels


def _align_progressive_return(
    progetto_voci: Sequence[VoceComputo],
    indice_ritorno: dict[str, list[dict[str, Any]]],
    ritorno_wrappers: Sequence[dict[str, Any]],
) -> _ReturnAlignmentResult:
    voci_allineate: list[ParsedVoce] = []
    legacy_pairs: list[tuple[VoceComputo | None, ParsedVoce]] = []
    price_adjustments: list[str] = []
    zero_guard_inputs: list[ParsedVoce] = []
    return_only_labels: list[str] = []
    progress_quantity_mismatches: list[str] = []
    progress_price_conflicts: list[str] = []
    matched_count = 0
    progress_price_registry: dict[tuple[int, str], float] = {}

    for voce_progetto in progetto_voci:
        keys = _keys_from_voce_progetto(voce_progetto)
        match = _pick_match(indice_ritorno, keys, voce_progetto)

        quantita = voce_progetto.quantita
        prezzo_unitario = voce_progetto.prezzo_unitario
        importo = voce_progetto.importo
        lock_price_override = False
        enforce_zero = _requires_zero_guard(
            voce_progetto.codice,
            voce_progetto.descrizione,
        )
        zero_guard_price_input: float | None = None
        zero_guard_quant_input: float | None = None
        zero_guard_import_input: float | None = None
        progress_key = _progress_price_key(voce_progetto)

        if match:
            matched_count += 1
            (
                price_from_match,
                quant_from_match,
                import_from_match,
            ) = _price_bundle(match)
            zero_guard_price_input = price_from_match
            zero_guard_quant_input = quant_from_match
            zero_guard_import_input = import_from_match
            if price_from_match is not None:
                corrected_price, was_corrected = _stabilize_return_price(
                    price_from_match,
                    voce_progetto.prezzo_unitario,
                )
                if was_corrected:
                    price_adjustments.append(
                        f"{_voce_label(voce_progetto)}: {format(price_from_match, '.2f')} -> {format(corrected_price, '.2f')}"
                    )
                    price_from_match = corrected_price

            project_quantity = voce_progetto.quantita
            if quant_from_match not in (None,) and project_quantity not in (None, 0):
                if not _quantities_match(project_quantity, quant_from_match):
                    progress_quantity_mismatches.append(
                        f"{_shorten_label(_voce_label(voce_progetto))} (ritorno="
                        f"{_format_quantity_for_warning(quant_from_match)} vs computo="
                        f"{_format_quantity_for_warning(project_quantity)})"
                    )
            # preferisci la quantità del ritorno, se presente; altrimenti usa quella di progetto
            if quant_from_match not in (None,):
                quantita = quant_from_match
            elif project_quantity is not None:
                quantita = project_quantity
        if price_from_match is not None:
            prezzo_unitario = round(price_from_match, 4)
            if quantita not in (None, 0):
                importo = round(prezzo_unitario * quantita, 2)
            if progress_key:
                existing_price = progress_price_registry.get(progress_key)
                if existing_price is None:
                    progress_price_registry[progress_key] = prezzo_unitario
                elif not _prices_match(existing_price, prezzo_unitario):
                    progress_price_conflicts.append(
                        f"{_shorten_label(_voce_label(voce_progetto))}"
                        f" ({format(existing_price, '.4f')} vs {format(prezzo_unitario, '.4f')})"
                    )
            elif import_from_match is not None:
                importo = _ceil_amount(import_from_match)
                if importo is not None and quantita not in (None, 0):
                    prezzo_unitario = round(importo / quantita, 4)
                else:
                    prezzo_unitario = prezzo_unitario

        if quantita is None and voce_progetto.quantita is not None:
            quantita = voce_progetto.quantita
        if quantita in (0, 0.0):
            importo = 0.0

        parsed_voce = _build_parsed_from_progetto(
            voce_progetto,
            quantita,
            prezzo_unitario,
            importo,
        )
        if match is None:
            parsed_voce.quantita = 0.0
            parsed_voce.importo = 0.0
            meta = dict(parsed_voce.metadata or {})
            meta["missing_from_return"] = True
            parsed_voce.metadata = meta

        if enforce_zero and match:
            zero_guard_inputs.append(
                _build_zero_guard_entry(
                    voce_progetto.codice or parsed_voce.codice,
                    voce_progetto.descrizione or parsed_voce.descrizione,
                    zero_guard_quant_input,
                    zero_guard_price_input,
                    zero_guard_import_input,
                )
            )

        voci_allineate.append(parsed_voce)
        legacy_pairs.append((voce_progetto, parsed_voce))

    return_only_labels = _collect_return_only_labels(ritorno_wrappers)
    return _ReturnAlignmentResult(
        voci_allineate=voci_allineate,
        legacy_pairs=legacy_pairs,
        matched_count=matched_count,
        price_adjustments=price_adjustments,
        zero_guard_inputs=zero_guard_inputs,
        return_only_labels=return_only_labels,
        progress_quantity_mismatches=progress_quantity_mismatches,
        progress_price_conflicts=progress_price_conflicts,
        excel_only_groups=[],
    )


def _align_totals_return(
    progetto_voci: Sequence[VoceComputo],
    indice_ritorno: dict[str, list[dict[str, Any]]],
    ritorno_wrappers: Sequence[dict[str, Any]],
    wbs_wrapper_map: dict[str, list[dict[str, Any]]],
    description_price_map: dict[str, list[float]],
    excel_group_targets: dict[str, Decimal],
    excel_group_labels: dict[str, str],
    excel_group_details: dict[str, dict[str, Any]],
) -> _ReturnAlignmentResult:
    voci_allineate: list[ParsedVoce] = []
    legacy_pairs: list[tuple[VoceComputo | None, ParsedVoce]] = []
    price_adjustments: list[str] = []
    zero_guard_inputs: list[ParsedVoce] = []
    group_entries: dict[str, list[ParsedVoce]] = {}
    project_group_entries: dict[str, list[tuple[ParsedVoce, VoceComputo]]] = {}
    project_primary_entries: dict[str, list[tuple[ParsedVoce, VoceComputo]]] = {}
    project_code_entries: dict[str, list[tuple[ParsedVoce, VoceComputo]]] = {}
    project_description_entries: dict[str, list[tuple[ParsedVoce, VoceComputo]]] = {}
    matched_count = 0
    wbs_price_map = _build_wbs_price_map(wbs_wrapper_map)
    used_wbs_keys: set[str] = set()

    for voce_progetto in progetto_voci:
        project_wbs_key = _wbs_key_from_model(voce_progetto)
        project_base_key = _base_wbs_key_from_key(project_wbs_key)
        keys = _keys_from_voce_progetto(voce_progetto)
        match = None
        quant_from_match: float | None = None
        import_from_match: float | None = None
        price_from_match: float | None = None
        matched_from_wbs = False
        matched_from_description = False
        wbs_group_key: str | None = None
        description_signature = _description_signature_from_model(voce_progetto)
        has_description_price = (
            description_signature is not None
            and description_signature in description_price_map
        )

        if project_base_key and project_base_key in wbs_price_map and not has_description_price:
            info = wbs_price_map[project_base_key]
            price_from_match = info.get("price")
            wbs_group_key = info.get("group_key")
            matched_from_wbs = True
            matched_count += 1
            used_wbs_keys.add(project_base_key)
        else:
            match = _pick_match(indice_ritorno, keys, voce_progetto)
            if match:
                matched_count += 1

        quantita = voce_progetto.quantita
        prezzo_unitario = voce_progetto.prezzo_unitario
        importo = voce_progetto.importo
        lock_price_override = False
        enforce_zero = _requires_zero_guard(
            voce_progetto.codice,
            voce_progetto.descrizione,
        )
        zero_guard_price_input: float | None = None
        zero_guard_quant_input: float | None = None
        zero_guard_import_input: float | None = None

        if match:
            (
                price_from_match_match,
                quant_from_match,
                import_from_match,
            ) = _price_bundle(match)
            price_from_match = price_from_match_match
            zero_guard_price_input = price_from_match
            zero_guard_quant_input = quant_from_match
            zero_guard_import_input = import_from_match
            if price_from_match is not None:
                corrected_price, was_corrected = _stabilize_return_price(
                    price_from_match,
                    voce_progetto.prezzo_unitario,
                )
                if was_corrected:
                    price_adjustments.append(
                        f"{_voce_label(voce_progetto)}: {format(price_from_match, '.2f')} -> {format(corrected_price, '.2f')}"
                    )
                    price_from_match = corrected_price
        elif matched_from_wbs and price_from_match is not None:
            zero_guard_price_input = price_from_match
            if voce_progetto.quantita not in (None, 0):
                zero_guard_quant_input = voce_progetto.quantita
                zero_guard_import_input = price_from_match * voce_progetto.quantita

        if price_from_match is None and description_signature:
            price_candidate = description_price_map.get(description_signature)
            if price_candidate is not None:
                price_from_match = price_candidate
                matched_from_description = True
        if price_from_match is not None:
            prezzo_unitario = price_from_match
            if quantita not in (None, 0):
                importo = price_from_match * quantita
            lock_price_override = True
        elif import_from_match is not None:
            importo = import_from_match
            if quantita not in (None, 0):
                prezzo_unitario = import_from_match / quantita
            else:
                prezzo_unitario = None
            lock_price_override = True

        if quantita is None and voce_progetto.quantita is not None:
            quantita = voce_progetto.quantita

        parsed_voce = _build_parsed_from_progetto(
            voce_progetto,
            quantita,
            prezzo_unitario,
            importo,
        )
        if lock_price_override:
            meta = dict(parsed_voce.metadata or {})
            meta["lock_return_price"] = True
            parsed_voce.metadata = meta
        if match is None and not (matched_from_wbs or matched_from_description):
            parsed_voce.quantita = 0.0
            parsed_voce.importo = 0.0
            meta = dict(parsed_voce.metadata or {})
            meta["missing_from_return"] = True
            parsed_voce.metadata = meta

        if enforce_zero and (match or matched_from_wbs or matched_from_description):
            zero_guard_inputs.append(
                _build_zero_guard_entry(
                    voce_progetto.codice or parsed_voce.codice,
                    voce_progetto.descrizione or parsed_voce.descrizione,
                    zero_guard_quant_input,
                    zero_guard_price_input,
                    zero_guard_import_input,
                )
            )

        group_key = _wbs_key_from_parsed(match) if match else None
        if not group_key and matched_from_wbs:
            group_key = wbs_group_key or project_wbs_key
        if not group_key and match:
            group_key = project_wbs_key
        if group_key and (match or matched_from_wbs):
            group_entries.setdefault(group_key, []).append(parsed_voce)

        project_key = project_wbs_key
        if project_key:
            project_group_entries.setdefault(project_key, []).append(
                (parsed_voce, voce_progetto)
            )
            primary, _ = _split_wbs_key(project_key)
            if primary:
                project_primary_entries.setdefault(primary, []).append(
                    (parsed_voce, voce_progetto)
                )
            for code_token in _collect_code_tokens(voce_progetto.codice):
                project_code_entries.setdefault(code_token, []).append(
                    (parsed_voce, voce_progetto)
                )
            for desc_token in _collect_description_tokens(voce_progetto.descrizione):
                project_description_entries.setdefault(desc_token, []).append(
                    (parsed_voce, voce_progetto)
                )

        voci_allineate.append(parsed_voce)
        legacy_pairs.append((voce_progetto, parsed_voce))

    resolved_from_groups, satisfied_group_keys = _distribute_group_targets(
        excel_group_targets,
        group_entries,
        project_group_entries,
        project_code_entries,
        project_primary_entries,
        project_description_entries,
        excel_group_details,
    )
    matched_count += resolved_from_groups

    return_only_labels = _collect_return_only_labels(
        ritorno_wrappers,
        satisfied_group_keys,
    )

    for key, entries in group_entries.items():
        target = excel_group_targets.get(key)
        if target is None:
            continue
        _apply_rounding_to_match(entries, target)

    excel_only_groups = [
        excel_group_labels[key]
        for key, total in excel_group_targets.items()
        if key not in group_entries and total != Decimal("0")
    ]

    for base_key in used_wbs_keys:
        for wrapper in wbs_wrapper_map.get(base_key, []):
            wrapper["matched"] = True

    return _ReturnAlignmentResult(
        voci_allineate=voci_allineate,
        legacy_pairs=legacy_pairs,
        matched_count=matched_count,
        price_adjustments=price_adjustments,
        zero_guard_inputs=zero_guard_inputs,
        return_only_labels=return_only_labels,
        progress_quantity_mismatches=[],
        progress_price_conflicts=[],
        excel_only_groups=excel_only_groups,
    )


def _align_return_rows(
    progetto_voci: Sequence[VoceComputo],
    ritorno_voci: Sequence[ParsedVoce],
    *,
    prefer_progressivi: bool,
    description_price_map: dict[str, list[float]],
) -> _ReturnAlignmentResult:
    indice_ritorno, ritorno_wrappers = _build_return_index(ritorno_voci)
    has_progressivi = prefer_progressivi and _has_progressivi(ritorno_voci)
    if has_progressivi:
        project_buckets = _build_project_description_buckets(progetto_voci)
        _assign_wrapper_preferences(ritorno_wrappers, project_buckets)
        return _align_progressive_return(
            progetto_voci,
            indice_ritorno,
            ritorno_wrappers,
        )
    return _align_description_only_return(
        progetto_voci,
        ritorno_voci,
        description_price_map,
    )


def _align_description_only_return(
    progetto_voci: Sequence[VoceComputo],
    ritorno_voci: Sequence[ParsedVoce],
    description_price_map: dict[str, list[float]],
) -> _ReturnAlignmentResult:
    excel_entries: list[dict[str, Any]] = []
    signature_queues: dict[str, deque[int]] = defaultdict(deque)
    signature_labels: dict[str, str] = {}

    for voce in ritorno_voci:
        signature = _description_signature_from_parsed(voce)
        if not signature:
            continue
        tokens = _descr_tokens(voce.descrizione)
        label = voce.descrizione or voce.codice or "voce senza descrizione"
        entry = {
            "signature": signature,
            "price": voce.prezzo_unitario or 0.0,
            "label": label,
            "tokens": tokens,
            "used": False,
        }
        idx = len(excel_entries)
        excel_entries.append(entry)
        signature_queues[signature].append(idx)
        signature_labels.setdefault(signature, label)

    signature_projects: dict[str, list[VoceComputo]] = defaultdict(list)
    for voce in progetto_voci:
        signature = _description_signature_from_model(voce)
        if signature:
            signature_projects[signature].append(voce)

    price_overrides: dict[int, tuple[float, int]] = {}
    matched_count = 0
    zero_guard_inputs: list[ParsedVoce] = []

    def _assign_from_signature(sig: str, voce: VoceComputo) -> bool:
        nonlocal matched_count
        queue = signature_queues.get(sig)
        if not queue:
            return False
        while queue:
            idx = queue.popleft()
            entry = excel_entries[idx]
            if entry["used"]:
                continue
            entry["used"] = True
            price_overrides[voce.id] = (entry["price"], idx)
            matched_count += 1
            return True
        return False

    def _project_sort_key(voce: VoceComputo) -> tuple[str, int]:
        code = voce.codice or ""
        order = voce.ordine or 0
        return (code, order)

    for sig, items in signature_projects.items():
        ordered = sorted(items, key=_project_sort_key)
        for voce in ordered:
            _assign_from_signature(sig, voce)

    unmatched_voice_ids = {voce.id for voce in progetto_voci if voce.id not in price_overrides}
    candidate_indices = [idx for idx, entry in enumerate(excel_entries) if not entry["used"]]
    for voce in progetto_voci:
        if voce.id not in unmatched_voice_ids:
            continue
        idx = _match_excel_entry_fuzzy(voce, excel_entries, candidate_indices)
        if idx is None:
            continue
        entry = excel_entries[idx]
        entry["used"] = True
        candidate_indices.remove(idx)
        price_overrides[voce.id] = (entry["price"], idx)
        matched_count += 1

    voci_allineate: list[ParsedVoce] = []
    legacy_pairs: list[tuple[VoceComputo | None, ParsedVoce]] = []

    for voce_progetto in progetto_voci:
        quantita = voce_progetto.quantita
        prezzo_unitario = voce_progetto.prezzo_unitario
        importo = voce_progetto.importo
        override = price_overrides.get(voce_progetto.id)
        lock_price_override = False

        if override is not None:
            price_from_excel, _ = override
            prezzo_unitario = price_from_excel
            if quantita not in (None, 0):
                importo = price_from_excel * quantita
            else:
                importo = None
            lock_price_override = True
            if _requires_zero_guard(voce_progetto.codice, voce_progetto.descrizione):
                zero_guard_inputs.append(
                    _build_zero_guard_entry(
                        voce_progetto.codice,
                        voce_progetto.descrizione,
                        quantita,
                        price_from_excel,
                        importo,
                    )
                )

        parsed_voce = _build_parsed_from_progetto(
            voce_progetto,
            quantita,
            prezzo_unitario,
            importo,
        )
        meta = dict(parsed_voce.metadata or {})
        if lock_price_override:
            meta["lock_return_price"] = True
        else:
            meta["missing_from_return"] = True
        parsed_voce.metadata = meta or None
        voci_allineate.append(parsed_voce)
        legacy_pairs.append((voce_progetto, parsed_voce))

    return_only_labels = [entry["label"] for entry in excel_entries if not entry["used"]]

    return _ReturnAlignmentResult(
        voci_allineate=voci_allineate,
        legacy_pairs=legacy_pairs,
        matched_count=matched_count,
        price_adjustments=[],
        zero_guard_inputs=zero_guard_inputs,
        return_only_labels=return_only_labels,
        progress_quantity_mismatches=[],
        progress_price_conflicts=[],
        excel_only_groups=[],
    )


class _WbsNormalizeContext:
    """Gestisce la creazione/ricerca dei nodi WBS e delle voci normalizzate."""

    def __init__(self, session: Session, commessa_id: int) -> None:
        self.session = session
        self.commessa_id = commessa_id
        self.spatial_cache: Dict[Tuple[int, int, str], WbsSpaziale] = {}
        self.wbs6_cache: Dict[str, Wbs6] = {}
        self.wbs7_cache: Dict[Tuple[int, Optional[str]], Wbs7] = {}
        self.voce_cache: Dict[Tuple[int, Optional[int], Optional[str], int], VoceNorm] = {}
        self.voce_by_legacy: Dict[int, VoceNorm] = {}
        self.impresa_cache: Dict[str, Impresa] = {}
        self._price_list_item_cache: Dict[str, Optional[int]] = {}

    def ensure_voce(
        self,
        parsed: ParsedVoce,
        legacy: Optional[VoceComputo],
        *,
        price_list_item_id: Optional[int] = None,
    ) -> Optional[VoceNorm]:
        info = self._analyze_parsed(parsed)
        if not info:
            return None
        spatial_levels, wbs6_code, wbs6_desc, wbs7_code, wbs7_desc = info
        spatial_leaf = self._ensure_spatial_hierarchy(spatial_levels)
        wbs6 = self._ensure_wbs6(wbs6_code, wbs6_desc, spatial_leaf)
        wbs7 = self._ensure_wbs7(wbs6, wbs7_code, wbs7_desc)
        resolved_price_list_id = price_list_item_id or self.resolve_price_list_item_id(parsed)

        target_wbs7_id = wbs7.id if wbs7 else None
        key = (wbs6.id, target_wbs7_id, parsed.codice, parsed.ordine)
        voce = self.voce_cache.get(key)
        if not voce and legacy:
            voce = self.get_voce_from_legacy(legacy.id)
        if not voce:
            voce = self.session.exec(
                select(VoceNorm).where(
                    VoceNorm.commessa_id == self.commessa_id,
                    VoceNorm.wbs6_id == wbs6.id,
                    VoceNorm.wbs7_id == target_wbs7_id,
                    VoceNorm.codice == parsed.codice,
                    VoceNorm.ordine == parsed.ordine,
                )
            ).first()
        if voce:
            updated = False
            if voce.wbs6_id != wbs6.id:
                voce.wbs6_id = wbs6.id
                updated = True
            if voce.wbs7_id != target_wbs7_id:
                voce.wbs7_id = target_wbs7_id
                updated = True
            if voce.codice != parsed.codice:
                voce.codice = parsed.codice
                updated = True
            if voce.ordine != parsed.ordine:
                voce.ordine = parsed.ordine
                updated = True
            if parsed.descrizione and voce.descrizione != parsed.descrizione:
                voce.descrizione = parsed.descrizione
                updated = True
            if parsed.unita_misura and voce.unita_misura != parsed.unita_misura:
                voce.unita_misura = parsed.unita_misura
                updated = True
            if parsed.note and voce.note != parsed.note:
                voce.note = parsed.note
                updated = True
            if legacy and voce.legacy_vocecomputo_id is None:
                voce.legacy_vocecomputo_id = legacy.id
                updated = True
            if resolved_price_list_id and voce.price_list_item_id is None:
                voce.price_list_item_id = resolved_price_list_id
                updated = True
            if updated:
                self.session.add(voce)
        else:
            voce = VoceNorm(
                commessa_id=self.commessa_id,
                wbs6_id=wbs6.id,
                wbs7_id=target_wbs7_id,
                codice=parsed.codice,
                descrizione=parsed.descrizione,
                unita_misura=parsed.unita_misura,
                note=parsed.note,
                ordine=parsed.ordine,
                legacy_vocecomputo_id=legacy.id if legacy else None,
                price_list_item_id=resolved_price_list_id,
            )
            self.session.add(voce)
            self.session.flush()
        self.voce_cache[key] = voce
        if voce.legacy_vocecomputo_id:
            self.voce_by_legacy[voce.legacy_vocecomputo_id] = voce
        return voce

    def get_voce_from_legacy(self, legacy_id: int) -> Optional[VoceNorm]:
        voce = self.voce_by_legacy.get(legacy_id)
        if voce:
            return voce
        voce = self.session.exec(
            select(VoceNorm).where(VoceNorm.legacy_vocecomputo_id == legacy_id)
        ).first()
        if voce:
            self.voce_by_legacy[legacy_id] = voce
            key = (voce.wbs6_id, voce.wbs7_id, voce.codice, voce.ordine)
            self.voce_cache[key] = voce
        return voce

    def get_or_create_impresa(self, label: Optional[str]) -> Optional[Impresa]:
        if not label:
            return None
        label = self._sanitize_impresa_label(label)
        if not label:
            return None
        normalized = re.sub(r"\s+", " ", label.strip()).lower()
        if not normalized:
            return None
        impresa = self.impresa_cache.get(normalized)
        if impresa:
            return impresa
        impresa = self.session.exec(
            select(Impresa).where(Impresa.normalized_label == normalized)
        ).first()
        if not impresa:
            impresa = Impresa(label=label.strip(), normalized_label=normalized)
            self.session.add(impresa)
            self.session.flush()
        self.impresa_cache[normalized] = impresa
        return impresa

    @staticmethod
    def _sanitize_impresa_label(label: str | None) -> str | None:
        if not label:
            return None
        text = label.strip()
        if not text:
            return None
        # Rimuove suffissi duplicati tipo "(2)" o "(3)"
        text = re.sub(r"\s*\(\d+\)\s*$", "", text).strip()
        return text or None

    def resolve_price_list_item_id(self, parsed: ParsedVoce) -> Optional[int]:
        metadata = parsed.metadata or {}
        product_id = metadata.get("product_id")
        if not product_id:
            return None
        cached = self._price_list_item_cache.get(product_id)
        if cached is not None:
            return cached
        item = self.session.exec(
            select(PriceListItem).where(
                PriceListItem.commessa_id == self.commessa_id,
                PriceListItem.product_id == product_id,
            )
        ).first()
        item_id = item.id if item else None
        self._price_list_item_cache[product_id] = item_id
        return item_id

    def _analyze_parsed(
        self,
        parsed: ParsedVoce,
    ) -> Optional[Tuple[list[ParsedWbsLevel], str, Optional[str], Optional[str], Optional[str]]]:
        spatial = [lvl for lvl in parsed.wbs_levels if lvl.level <= 5]
        wbs6_level = next((lvl for lvl in parsed.wbs_levels if lvl.level == 6), None)
        wbs7_level = next((lvl for lvl in parsed.wbs_levels if lvl.level == 7), None)
        wbs6_code = _normalize_wbs6_code(wbs6_level.code if wbs6_level else None)
        if not wbs6_code:
            return None
        wbs6_desc = (wbs6_level.description if wbs6_level else None) or f"WBS6 {wbs6_code}"
        wbs7_code = _normalize_wbs7_code(wbs7_level.code if wbs7_level else None)
        wbs7_desc = wbs7_level.description if wbs7_level else None
        return spatial, wbs6_code, wbs6_desc, wbs7_code, wbs7_desc

    def _ensure_spatial_hierarchy(self, levels: Sequence[ParsedWbsLevel]) -> Optional[WbsSpaziale]:
        parent: Optional[WbsSpaziale] = None
        last: Optional[WbsSpaziale] = None
        for lvl in levels:
            code = (lvl.code or "").strip()
            if not code:
                continue
            key = (self.commessa_id, lvl.level, code)
            node = self.spatial_cache.get(key)
            if not node:
                node = self.session.exec(
                    select(WbsSpaziale).where(
                        WbsSpaziale.commessa_id == self.commessa_id,
                        WbsSpaziale.level == lvl.level,
                        WbsSpaziale.code == code,
                    )
                ).first()
            if not node:
                node = WbsSpaziale(
                    commessa_id=self.commessa_id,
                    parent_id=parent.id if parent else None,
                    level=lvl.level,
                    code=code,
                    description=lvl.description,
                )
                self.session.add(node)
                self.session.flush()
            self.spatial_cache[key] = node
            parent = node
            last = node
        return last

    def _ensure_wbs6(
        self,
        code: str,
        description: Optional[str],
        spatial_leaf: Optional[WbsSpaziale],
    ) -> Wbs6:
        node = self.wbs6_cache.get(code)
        if not node:
            node = self.session.exec(
                select(Wbs6).where(
                    Wbs6.commessa_id == self.commessa_id,
                    Wbs6.code == code,
                )
            ).first()
        if not node:
            desc = description or f"WBS6 {code}"
            label = f"{code} - {desc}" if desc else code
            node = Wbs6(
                commessa_id=self.commessa_id,
                wbs_spaziale_id=spatial_leaf.id if spatial_leaf else None,
                code=code,
                description=desc,
                label=label,
            )
            self.session.add(node)
            self.session.flush()
        elif spatial_leaf and node.wbs_spaziale_id is None:
            node.wbs_spaziale_id = spatial_leaf.id
            self.session.add(node)
        self.wbs6_cache[code] = node
        return node

    def _ensure_wbs7(
        self,
        wbs6: Wbs6,
        code: Optional[str],
        description: Optional[str],
    ) -> Optional[Wbs7]:
        if not code:
            return None
        key = (wbs6.id, code)
        node = self.wbs7_cache.get(key)
        if not node:
            node = self.session.exec(
                select(Wbs7).where(
                    Wbs7.wbs6_id == wbs6.id,
                    Wbs7.commessa_id == self.commessa_id,
                    Wbs7.code == code,
                )
            ).first()
        if not node:
            node = Wbs7(
                commessa_id=self.commessa_id,
                wbs6_id=wbs6.id,
                code=code,
                description=description,
            )
            self.session.add(node)
            self.session.flush()
        self.wbs7_cache[key] = node
        return node


def _parse_custom_return_excel(
    file_path: Path,
    sheet_name: str | None,
    code_columns: Sequence[str],
    description_columns: Sequence[str],
    price_column: str,
    quantity_column: str | None = None,
    progressive_column: str | None = None,
) -> ParsedComputo:
    workbook = load_workbook(
        filename=file_path,
        data_only=True,
        read_only=True,
    )
    try:
        sheet = _select_sheet(workbook, sheet_name)
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    header_idx = _locate_header_row(rows)
    if header_idx is None:
        raise ValueError("Il foglio Excel selezionato non contiene righe valide da importare")
    data_rows = rows[header_idx + 1 :]
    if not data_rows:
        raise ValueError("Il foglio Excel selezionato non contiene dati dopo l'intestazione")

    header_row = rows[header_idx]
    detection = _detect_column_suggestions(rows, header_idx)
    suggestions: dict[str, _ColumnSuggestion] = detection.get("suggestions") if detection else {}
    profiles: list[_ColumnProfile] = detection.get("profiles") if detection else []

    def _format_profiles() -> str:
        entries: list[str] = []
        for profile in profiles[:10]:
            label = profile.header_label or "—"
            entries.append(f"{profile.letter}: {label}")
        return ", ".join(entries)

    def _warn_and_use(target: str, suggestion: _ColumnSuggestion | None, warnings: list[str]) -> int | None:
        if suggestion is None:
            return None
        label = suggestion.header_label or suggestion.column_letter
        warnings.append(
            f"Colonna {target} non trovata nella configurazione salvata. "
            f"Utilizzo automaticamente '{label}' ({suggestion.column_letter})."
        )
        return suggestion.column_index

    def _ensure_indexes(name: str, indexes: list[int], warnings: list[str]) -> list[int]:
        if indexes and _column_has_values(data_rows, indexes):
            return indexes
        suggestion = suggestions.get(name)
        fallback = _warn_and_use(name, suggestion, warnings)
        if fallback is not None:
            return [fallback]
        available = _format_profiles()
        raise ValueError(
            f"Impossibile individuare la colonna {name}. "
            f"Intestazioni rilevate: {available or 'nessuna intestazione valida.'}"
        )

    column_warnings: list[str] = []
    try:
        code_indexes = _columns_to_indexes(code_columns, "codice", header_row=header_row, required=False)
    except ValueError:
        code_indexes = []
    try:
        description_indexes = _columns_to_indexes(description_columns, "descrizione", header_row=header_row, required=False)
    except ValueError:
        description_indexes = []
    try:
        price_index = _single_column_index(price_column, "prezzo unitario", header_row=header_row)
    except ValueError:
        price_index = None
    try:
        quantity_index = _single_column_index(quantity_column, "quantita", header_row=header_row, required=False)
    except ValueError:
        quantity_index = None
    try:
        progressive_index = _single_column_index(progressive_column, "progressivo", header_row=header_row, required=False)
    except ValueError:
        progressive_index = None

    code_indexes = _ensure_indexes("codice", code_indexes, column_warnings)
    description_indexes = _ensure_indexes("descrizione", description_indexes, column_warnings)
    if price_index is None:
        price_index = _warn_and_use("prezzo", suggestions.get("prezzo"), column_warnings)
    if price_index is None:
        raise ValueError(
            f"Seleziona la colonna da utilizzare per il campo prezzo unitario. "
            f"Intestazioni rilevate: {_format_profiles() or 'nessuna'}"
        )
    if quantity_index is None:
        quantity_index = _warn_and_use("quantità", suggestions.get("quantita"), column_warnings)
    if progressive_index is None:
        progressive_index = _warn_and_use("progressivo", suggestions.get("progressivo"), column_warnings)
    if not code_indexes and not description_indexes and progressive_index is None:
        raise ValueError(
            "Seleziona almeno una colonna da utilizzare come codice, descrizione o progressivo"
        )
    workbook_formulas = load_workbook(
        filename=file_path,
        data_only=False,
        read_only=True,
    )
    try:
        formula_sheet = _select_sheet(workbook_formulas, sheet_name)
        formula_rows = list(
            formula_sheet.iter_rows(
                min_row=header_idx + 2,
                values_only=False,
            )
        )
    finally:
        workbook_formulas.close()

    formula_rows_iter = iter(formula_rows)

    voci: list[ParsedVoce] = []
    ordine = 0
    treat_price_as_total = False
    for row in data_rows:
        formula_row = next(formula_rows_iter, ())
        if not _row_has_values(row):
            continue
        codice = _combine_text(row, code_indexes)
        descrizione = _combine_text(row, description_indexes)
        raw_price = _cell_to_float(row, price_index)
        quantita = _cell_to_float(row, quantity_index) if quantity_index is not None else None
        progressivo_value = _cell_to_progressive(row, progressive_index)
        if raw_price is None:
            raw_price = 0.0
        if not codice and not descrizione and progressivo_value is None:
            continue
        if not codice and progressivo_value is not None:
            codice = f"PROG-{progressivo_value:05d}"
        if not descrizione and progressivo_value is not None:
            descrizione = f"Voce progressivo {progressivo_value}"
        formula_cell = (
            formula_row[price_index] if price_index < len(formula_row) else None
        )
        if _has_external_formula(formula_cell):
            raise ValueError(
                "La colonna prezzo contiene formule collegate a file esterni. Apri il file in Excel e incolla i valori numerici prima di importare."
            )
        if quantity_index is not None and quantity_index < len(formula_row):
            quantity_formula_cell = formula_row[quantity_index]
        else:
            quantity_formula_cell = None
        if quantity_index is not None and _has_external_formula(quantity_formula_cell):
            raise ValueError(
                "La colonna quantità contiene formule collegate a file esterni. Incolla i valori numerici prima dell'import."
            )
        prezzo = _sanitize_price_candidate(raw_price)
        if prezzo is None:
            continue
        if treat_price_as_total:
            quantita_value = None
            prezzo_value: float | None = None
            importo_value = round(prezzo, 2)
        else:
            quantita_value = quantita if quantita not in (None,) else None
            prezzo_value = round(prezzo, 4)
            importo_value = (
                round(prezzo_value * quantita_value, 2)
                if quantita_value not in (None, 0)
                else None
            )
        voce_descrizione = descrizione or codice or "Voce senza descrizione"
        wbs_levels: list[ParsedWbsLevel] = []
        normalized_code_value = _normalize_wbs7_code(codice)
        if _looks_like_wbs7_code(normalized_code_value):
            wbs_levels.append(
                ParsedWbsLevel(
                    level=7,
                    code=normalized_code_value,
                    description=normalized_code_value,
                )
            )

        voci.append(
            ParsedVoce(
                ordine=ordine,
                progressivo=progressivo_value,
                codice=codice,
                descrizione=voce_descrizione,
                wbs_levels=wbs_levels,
                unita_misura=None,
                quantita=quantita_value,
                prezzo_unitario=prezzo_value,
                importo=importo_value,
                note=None,
                metadata={
                    "source": "excel_custom_return",
                    "sheet_name": sheet_name,
                    "code_columns": code_columns,
                    "description_columns": description_columns,
                    "price_column": price_column,
                    "quantity_column": quantity_column,
                    "progressive_column": progressive_column,
                    "group_total_only": treat_price_as_total,
                },
            )
        )
        ordine += 1

    if not voci:
        raise ValueError(
            "Nessuna voce valida trovata nel foglio selezionato. "
            "Verifica di aver scelto le colonne corrette e che contengano valori numerici per i prezzi."
        )

    titolo = f"Ritorno {sheet_name}" if sheet_name else "Ritorno"
    quantita_values = [voce.quantita for voce in voci if voce.quantita is not None]
    totale_quantita = (
        round(sum(quantita_values), 4) if quantita_values else None
    )
    parsed = ParsedComputo(
        titolo=titolo,
        totale_importo=None,
        totale_quantita=totale_quantita,
        voci=voci,
    )
    return _CustomReturnParseResult(
        computo=parsed,
        column_warnings=column_warnings,
    )


def _select_sheet(workbook, requested_name: str | None):
    if requested_name:
        for name in workbook.sheetnames:
            if name.lower() == requested_name.lower():
                return workbook[name]
        raise ValueError(f"Il foglio '{requested_name}' non è presente nel file Excel caricato")

    def _has_codice_header(sheet) -> bool:
        # cerca una riga che contenga la parola CODICE (o simile) nelle prime 50 righe
        for row in sheet.iter_rows(max_row=50, values_only=True):
            if not row:
                continue
            for cell in row:
                if isinstance(cell, str) and "codice" in cell.strip().lower():
                    return True
        return False

    for name in workbook.sheetnames:
        sheet = workbook[name]
        if _has_codice_header(sheet):
            return sheet

    return workbook.active


def _rows_to_dataframe(rows: Sequence[Sequence[Any]]) -> pd.DataFrame:
    if not rows:
        return pd.DataFrame()
    max_len = max(len(row) for row in rows)
    if max_len == 0:
        return pd.DataFrame()
    normalized_rows = []
    for row in rows:
        normalized = list(row) + [None] * (max_len - len(row))
        normalized_rows.append(normalized)
    columns = [get_column_letter(idx + 1) for idx in range(max_len)]
    return pd.DataFrame(normalized_rows, columns=columns)


def _ratio(values: Sequence[str], predicate) -> float:
    if not values:
        return 0.0
    count = sum(1 for value in values if predicate(value))
    return count / len(values)


def _looks_numeric(value: str) -> bool:
    text = value.strip().replace(".", "").replace("\u20AC", "").replace("%", "")
    text = text.replace(",", ".")
    try:
        float(text)
        return True
    except ValueError:
        return False


def _looks_currency(value: str) -> bool:
    text = value.strip().lower()
    has_digits = any(ch.isdigit() for ch in text)
    return "\u20ac" in text or "eur" in text or (has_digits and ("," in text or "." in text))


def _looks_code(value: str) -> bool:
    if not value:
        return False
    text = value.strip()
    if not text:
        return False
    return bool(_CODE_REGEX.match(text))


def _looks_text(value: str) -> bool:
    cleaned = re.sub(r"\d+", "", value or "").strip()
    return len(cleaned) >= 10


def _pick_column_profile(
    target: str,
    profiles: Sequence[_ColumnProfile],
    *,
    keywords: Sequence[str],
    prefer_numeric: bool = False,
    prefer_currency: bool = False,
    prefer_code: bool = False,
    prefer_text: bool = False,
) -> _ColumnSuggestion | None:
    best: _ColumnProfile | None = None
    best_score = 0.0
    for profile in profiles:
        header = (profile.header_label or "").lower()
        score = 0.0
        for keyword in keywords:
            if keyword and keyword.lower() in header:
                score += 5
        if prefer_numeric:
            score += profile.numeric_ratio * 3
        if prefer_currency:
            score += profile.currency_ratio * 4
        if prefer_code:
            score += profile.code_ratio * 4
        if prefer_text:
            score += profile.text_ratio * 2
        if profile.header_label:
            score += 0.5
        if score > best_score:
            best = profile
            best_score = score
    if best and best_score >= 1.0:
        return _ColumnSuggestion(
            target=target,
            column_index=best.index,
            column_letter=best.letter,
            header_label=best.header_label,
            score=best_score,
        )
    return None


def _detect_column_suggestions(
    rows: Sequence[Sequence[Any]],
    header_idx: int | None,
) -> dict[str, Any]:
    if header_idx is None:
        return {}
    df = _rows_to_dataframe(rows)
    if df.empty or header_idx >= len(df):
        return {}
    header_series = df.iloc[header_idx]
    samples_df = df.iloc[header_idx + 1 : header_idx + 11]
    profiles: list[_ColumnProfile] = []
    for idx, letter in enumerate(df.columns):
        header_value = header_series.get(letter)
        header_text = None
        if header_value is not None:
            text = str(header_value).strip()
            header_text = text or None
        column_samples = samples_df[letter]
        sample_values = [
            str(value).strip()
            for value in column_samples
            if value is not None and str(value).strip()
        ][:5]
        profiles.append(
            _ColumnProfile(
                index=idx,
                letter=letter,
                header_label=header_text,
                samples=sample_values,
                numeric_ratio=_ratio(sample_values, _looks_numeric),
                currency_ratio=_ratio(sample_values, _looks_currency),
                code_ratio=_ratio(sample_values, _looks_code),
                text_ratio=_ratio(sample_values, _looks_text),
            )
        )
    suggestions = {
        "codice": _pick_column_profile(
            "codice", profiles, keywords=["cod", "voce", "item", "voce computo"], prefer_code=True, prefer_text=False
        ),
        "descrizione": _pick_column_profile(
            "descrizione", profiles, keywords=["descr", "indicazione", "lavori"], prefer_text=True
        ),
        "prezzo": _pick_column_profile(
            "prezzo", profiles, keywords=["prezzo", "unit", "importo", "€/u"], prefer_numeric=True, prefer_currency=True
        ),
        "quantita": _pick_column_profile(
            "quantita", profiles, keywords=["quant", "q.t", "qta"], prefer_numeric=True
        ),
        "progressivo": _pick_column_profile(
            "progressivo", profiles, keywords=["prog", "progressivo", "n."], prefer_numeric=True
        ),
    }
    return {
        "suggestions": {key: value for key, value in suggestions.items() if value is not None},
        "profiles": profiles,
    }


def _locate_header_row(rows: list) -> int | None:
    best_idx: int | None = None
    best_score = 0.0
    for idx, row in enumerate(rows[:200]):
        values = [cell for cell in row if _cell_has_content(cell)]
        if not values:
            continue
        short_tokens = 0
        keyword_bonus = 0
        penalty = 0
        for cell in values:
            text = _cell_to_text(cell)
            if not text:
                continue
            if len(text) > 90:
                penalty += 1
                continue
            short_tokens += 1
            lowered = text.lower()
            if any(keyword in lowered for keyword in ("cod", "descr", "prezzo", "importo", "unit", "um", "qta", "quant")):
                keyword_bonus += 1
        if short_tokens == 0:
            continue
        density = short_tokens / max(len(row), 1)
        score = short_tokens + keyword_bonus * 1.5 + density - penalty * 0.5
        if idx <= 2:
            score -= 0.5
        if score > best_score:
            best_score = score
            best_idx = idx
    return best_idx


def _columns_to_indexes(
    columns: Sequence[str],
    label: str,
    *,
    header_row: Sequence[Any] | None = None,
    required: bool = True,
) -> list[int]:
    if not columns:
        if required:
            raise ValueError(f"Seleziona almeno una colonna per il campo {label}")
        return []
    indexes: list[int] = []
    for column in columns:
        normalized = (column or "").strip().lstrip("$")
        if not normalized:
            continue
        idx = _resolve_column_reference(normalized, header_row)
        if idx < 0:
            raise ValueError(f"Colonna '{column}' non valida per il campo {label}")
        if idx not in indexes:
            indexes.append(idx)
    if not indexes and required:
        raise ValueError(f"Seleziona almeno una colonna valida per il campo {label}")
    return indexes


def _single_column_index(
    column: str | None,
    label: str,
    *,
    header_row: Sequence[Any] | None = None,
    required: bool = True,
) -> int | None:
    normalized = (column or "").strip().lstrip("$")
    if not normalized:
        if required:
            raise ValueError(f"Seleziona la colonna da utilizzare per il campo {label}")
        return None
    idx = _resolve_column_reference(normalized, header_row)
    if idx < 0:
        raise ValueError(f"Colonna '{column}' non valida per il campo {label}")
    return idx


def _resolve_column_reference(
    reference: str,
    header_row: Sequence[Any] | None,
) -> int:
    if not reference:
        raise ValueError("Riferimento colonna non valido")
    try:
        return column_index_from_string(reference.upper()) - 1
    except ValueError:
        pass
    if header_row is None:
        raise ValueError(f"Colonna '{reference}' non valida")
    target = reference.strip().lower()
    for idx, value in enumerate(header_row):
        if value is None:
            continue
        text = str(value).strip().lower()
        if text == target:
            return idx
    raise ValueError(f"Colonna '{reference}' non valida")


def _normalize_header_text(value) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    return text


def _row_has_values(row) -> bool:
    for cell in row:
        if _cell_has_content(cell):
            return True
    return False


def _column_has_values(rows: Sequence[Sequence[Any]], indexes: Sequence[int]) -> bool:
    if not rows or not indexes:
        return False
    for row in rows:
        for idx in indexes:
            if idx < len(row) and _cell_has_content(row[idx]):
                return True
    return False


def _sanitize_price_candidate(value: float) -> float:
    """
    Restituisce il prezzo senza alterarne l'ordine di grandezza.

    In passato dividevamo automaticamente i valori superiori a 1.000 per evitare
    di confondere importi totali con prezzi unitari. Questo però falsava i listini
    reali (es. 2.200 € diventava 220 €) e innescava differenze del 90% nella
    dashboard. Ora ci limitiamo ad arrotondare alla quarta cifra decimale e, solo
    in caso di numeri veramente anomali (> 1.000.000), riduciamo progressivamente
    mantenendo comunque un ordine di grandezza plausibile.
    """
    abs_value = abs(value)
    if abs_value == 0:
        return 0.0

    if abs_value <= 1_000_000:
        return round(float(value), 4)

    normalized = float(value)
    while abs(normalized) > 1_000_000:
        normalized /= 10
    return round(normalized, 4)


def _has_external_formula(cell) -> bool:
    """
    Verifica se una cella contiene una formula collegata a file esterni.
    """
    if cell is None or getattr(cell, "data_type", None) != "f":
        return False
    formula = str(cell.value or "")
    return "[" in formula


def _cell_has_content(value) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        return bool(value.strip())
    return True


def _combine_text(row, indexes: Sequence[int]) -> str | None:
    parts: list[str] = []
    for idx in indexes:
        if idx < len(row):
            text = _cell_to_text(row[idx])
            if text:
                parts.append(text)
    if not parts:
        return None
    return " ".join(parts).strip()


def _cell_to_text(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        text = value.replace("_x000D_", " ").replace("\r", " ").replace("\n", " ").strip()
    else:
        text = str(value).replace("_x000D_", " ").replace("\r", " ").replace("\n", " ").strip()
    return text or None


def _cell_to_float(row, index: int) -> float | None:
    if index >= len(row):
        return None
    value = row[index]
    if value is None:
        return None
    if isinstance(value, (int, float, Decimal)):
        return float(value)
    text = str(value).replace("\u00a0", " ").strip()
    if not text:
        return None
    normalized = text
    for token in ("\u20ac", "€", "eur", "EUR", "Eur", "euro", "EURO"):
        normalized = normalized.replace(token, "")
    normalized = normalized.replace("%", "")
    normalized = "".join(ch for ch in normalized if not ch.isspace())
    if "," in normalized and "." in normalized:
        if normalized.rfind(",") > normalized.rfind("."):
            normalized = normalized.replace(".", "").replace(",", ".")
        else:
            normalized = normalized.replace(",", "")
    elif "," in normalized:
        normalized = normalized.replace(",", ".")
    try:
        return float(normalized)
    except ValueError:
        return None


def _cell_to_progressive(row, index: int | None) -> int | None:
    if index is None or index >= len(row):
        return None
    value = row[index]
    if value in (None, "", " "):
        return None
    if isinstance(value, (int, float, Decimal)):
        try:
            return int(float(value))
        except (TypeError, ValueError):
            return None
    text = str(value).strip()
    if not text:
        return None
    text = text.replace(",", ".")
    try:
        return int(float(text))
    except ValueError:
        return None


def _normalize_wbs6_code(value: Optional[str]) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    return text.upper() or None


def _normalize_wbs7_code(value: Optional[str]) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    return text.upper() or None


def _looks_like_wbs7_code(value: Optional[str]) -> bool:
    if not value:
        return False
    text = str(value).strip().upper()
    if "." not in text:
        return False
    if len(text) < 5:
        return False
    return bool(re.match(r"^[A-Z0-9]{1,4}(?:\.[A-Z0-9]{2,4})+$", text))


def _map_wbs_levels(levels: Sequence) -> dict[str, str | None]:
    data: dict[str, str | None] = {}
    by_level = {level.level: level for level in levels}
    for idx in range(1, MAX_WBS_LEVELS + 1):
        entry = by_level.get(idx)
        data[f"wbs_{idx}_code"] = entry.code if entry else None
        data[f"wbs_{idx}_description"] = entry.description if entry else None
    return data


def _normalize_commessa_tag(commessa_id: int | None, commessa_code: str | None) -> str | None:
    code = (commessa_code or "commessa").strip() or "commessa"
    identifier = commessa_id or 0
    return f"{identifier}::{code}"


def _build_global_voce_code(commessa_tag: str | None, parsed: ParsedVoce) -> str | None:
    if not commessa_tag:
        return None
    base = parsed.codice or (
        f"PROG-{parsed.progressivo}"
        if parsed.progressivo is not None
        else f"ORD-{parsed.ordine}"
    )
    if not base:
        return None
    normalized_code = str(base).strip()
    if not normalized_code:
        return None
    wbs6_code = None
    for livello in parsed.wbs_levels:
        if getattr(livello, "level", None) == 6 and getattr(livello, "code", None):
            wbs6_code = livello.code
            break
    suffix = normalized_code
    if wbs6_code:
        suffix = f"{suffix}@{wbs6_code}"
    return f"{commessa_tag}::{suffix}"


def _build_parsed_from_progetto(
    voce: VoceComputo,
    quantita: float | None,
    prezzo_unitario: float | None,
    importo: float | None,
) -> ParsedVoce:
    livelli: list[ParsedWbsLevel] = []
    for level in range(1, MAX_WBS_LEVELS + 1):
        code = getattr(voce, f"wbs_{level}_code")
        description = getattr(voce, f"wbs_{level}_description")
        if code or description:
            livelli.append(
                ParsedWbsLevel(level=level, code=code, description=description)
            )

    if quantita is not None:
        quantita = _ceil_quantity(quantita)

    if prezzo_unitario is not None:
        prezzo_unitario = round(float(prezzo_unitario), 4)
    if importo is not None:
        importo = _ceil_amount(importo)

    return ParsedVoce(
        ordine=voce.ordine,
        progressivo=voce.progressivo,
        codice=voce.codice,
        descrizione=voce.descrizione,
        wbs_levels=livelli,
        unita_misura=voce.unita_misura,
        quantita=quantita,
        prezzo_unitario=prezzo_unitario,
        importo=importo,
        note=voce.note,
        metadata=voce.extra_metadata,
    )


def _build_project_snapshot_from_price_offers(
    progetto_voci: Sequence[VoceComputo],
    price_items: Sequence[PriceListItem],
    offer_price_map: dict[int, float],
) -> tuple[list[ParsedVoce], list[tuple[VoceComputo, ParsedVoce]]]:
    if not progetto_voci:
        return [], []

    product_index: dict[str, PriceListItem] = {}
    for item in price_items:
        product_id = (item.product_id or "").strip()
        if product_id:
            product_index[product_id] = item

    parsed_voci: list[ParsedVoce] = []
    legacy_pairs: list[tuple[VoceComputo, ParsedVoce]] = []
    for voce in progetto_voci:
        metadata = voce.extra_metadata or {}
        product_id = metadata.get("product_id") if isinstance(metadata, dict) else None
        target_item = product_index.get(product_id) if product_id else None
        prezzo = offer_price_map.get(target_item.id) if target_item else None

        quantita = voce.quantita
        prezzo_value = prezzo if prezzo is not None else voce.prezzo_unitario
        importo_value = voce.importo
        if (
            prezzo is not None
            and quantita not in (None, 0)
        ):
            importo_value = _ceil_amount(
                Decimal(str(prezzo)) * Decimal(str(quantita))
            )

        parsed = _build_parsed_from_progetto(
            voce,
            quantita,
            prezzo_value,
            importo_value,
        )
        parsed_voci.append(parsed)
        legacy_pairs.append((voce, parsed))
    return parsed_voci, legacy_pairs


def _build_return_index(
    voci: Sequence[ParsedVoce],
) -> tuple[dict[str, list[dict[str, Any]]], list[dict[str, Any]]]:
    index: dict[str, list[dict[str, Any]]] = {}
    wrappers: list[dict[str, Any]] = []
    for voce in voci:
        base_key = _wbs_base_key_from_parsed(voce)
        full_key = _wbs_key_from_parsed(voce)
        is_wbs_shareable = (
            base_key is not None
            and voce.progressivo is None
            and _looks_like_wbs7_code(voce.codice)
        )
        wrapper: dict[str, Any] = {
            "voce": voce,
            "used": False,
            "matched": False,
            "tokens": _descr_tokens(voce.descrizione),
            "base_key": base_key,
            "full_key": full_key,
            "reusable": is_wbs_shareable,
        }
        wrappers.append(wrapper)
        keys = _keys_from_parsed_voce(voce)
        if not keys:
            continue
        for key in keys:
            index.setdefault(key, []).append(wrapper)
        if base_key:
            index.setdefault(f"__wbs__:{base_key}", []).append(wrapper)
    return index, wrappers


def _build_wbs_wrapper_map(
    wrappers: Sequence[dict[str, Any]]
) -> dict[str, list[dict[str, Any]]]:
    mapping: dict[str, list[dict[str, Any]]] = {}
    for wrapper in wrappers:
        base_key = wrapper.get("base_key")
        if not base_key:
            continue
        mapping.setdefault(base_key, []).append(wrapper)
    return mapping


def _build_wbs_price_map(
    wbs_wrapper_map: dict[str, list[dict[str, Any]]],
) -> dict[str, dict[str, Any]]:
    price_map: dict[str, dict[str, Any]] = {}
    for base_key, wrappers in wbs_wrapper_map.items():
        for wrapper in wrappers:
            if not wrapper.get("reusable"):
                continue
            voce = wrapper["voce"]
            price = voce.prezzo_unitario
            if price in (None, 0):
                continue
            price_map[base_key] = {
                "price": price,
                "group_key": wrapper.get("full_key") or _wbs_key_from_parsed(voce),
            }
            break
    return price_map


def _build_description_price_map(
    ritorno_voci: Sequence[ParsedVoce],
) -> dict[str, list[float]]:
    mapping: dict[str, list[float]] = defaultdict(list)
    for voce in ritorno_voci:
        signature = _description_signature_from_parsed(voce)
        if not signature:
            continue
        mapping[signature].append(voce.prezzo_unitario or 0.0)
    return dict(mapping)


def _build_price_list_lookup(
    items: Sequence[PriceListItem],
) -> tuple[
    dict[str, list[PriceListItem]],
    dict[str, list[PriceListItem]],
    dict[str, list[PriceListItem]],
    dict[str, list[PriceListItem]],
    dict[str, list[PriceListItem]],
    dict[str, list[tuple[PriceListItem, list[float]]]],
]:
    code_map: dict[str, list[PriceListItem]] = defaultdict(list)
    signature_map: dict[str, list[PriceListItem]] = defaultdict(list)
    description_map: dict[str, list[PriceListItem]] = defaultdict(list)
    head_signature_map: dict[str, list[PriceListItem]] = defaultdict(list)
    tail_signature_map: dict[str, list[PriceListItem]] = defaultdict(list)
    embedding_map: dict[str, list[tuple[PriceListItem, list[float]]]] = defaultdict(list)

    for item in items:
        code_token = _normalize_code_token(item.item_code)
        if code_token:
            code_map[code_token].append(item)

        signature = _description_signature(
            item.item_description,
            item.unit_label,
            item.wbs6_code,
        )
        if signature:
            signature_map[signature].append(item)

        desc_token = _normalize_description_token(item.item_description)
        if desc_token:
            description_map[desc_token].append(item)

        head_sig, tail_sig = _head_tail_signatures(item.item_description)
        if head_sig:
            head_signature_map[head_sig].append(item)
        if tail_sig:
            tail_signature_map[tail_sig].append(item)

        metadata = item.extra_metadata or {}
        if isinstance(metadata, dict):
            nlp_payload = metadata.get("nlp")
            if isinstance(nlp_payload, dict):
                embedding_info = semantic_embedding_service.extract_embedding_payload(nlp_payload)
                if isinstance(embedding_info, dict):
                    vector = embedding_info.get("vector")
                    model_id = embedding_info.get("model_id") or nlp_payload.get("model_id")
                    if model_id and model_id != semantic_embedding_service.model_id:
                        continue
                    if isinstance(vector, list) and vector:
                        normalized_wbs6 = _normalize_code_token(item.wbs6_code) or _SEMANTIC_DEFAULT_BUCKET
                        payload = (item, [float(val) for val in vector])
                        embedding_map[normalized_wbs6].append(payload)
                        embedding_map[_SEMANTIC_DEFAULT_BUCKET].append(payload)

    return (
        code_map,
        signature_map,
        description_map,
        head_signature_map,
        tail_signature_map,
        embedding_map,
    )


def _match_price_list_item_entry(
    parsed: ParsedVoce,
    code_map: dict[str, list[PriceListItem]],
    signature_map: dict[str, list[PriceListItem]],
    description_map: dict[str, list[PriceListItem]],
    head_signature_map: dict[str, list[PriceListItem]],
    tail_signature_map: dict[str, list[PriceListItem]],
    embedding_map: dict[str, list[tuple[PriceListItem, list[float]]]],
) -> PriceListItem | None:
    code_token = _normalize_code_token(parsed.codice)
    if code_token:
        candidates = code_map.get(code_token, [])
        candidate = _select_price_list_item_candidate(candidates, parsed)
        if candidate:
            return candidate

    signature = _description_signature_from_parsed(parsed)
    if signature:
        candidate = _select_price_list_item_candidate(
            signature_map.get(signature, []),
            parsed,
        )
        if candidate:
            return candidate

    desc_token = _normalize_description_token(parsed.descrizione)
    if desc_token:
        return _select_price_list_item_candidate(
            description_map.get(desc_token, []),
            parsed,
        )
    head_sig, tail_sig = _head_tail_signatures(parsed.descrizione)
    if head_sig:
        candidate = _select_price_list_item_candidate(
            head_signature_map.get(head_sig, []),
            parsed,
        )
        if candidate:
            return candidate
    if tail_sig:
        candidate = _select_price_list_item_candidate(
            tail_signature_map.get(tail_sig, []),
            parsed,
        )
        if candidate:
            return candidate
    return _match_price_list_item_semantic(parsed, embedding_map)


def _select_price_list_item_candidate(
    candidates: Sequence[PriceListItem],
    parsed: ParsedVoce,
) -> PriceListItem | None:
    if not candidates:
        return None
    if len(candidates) == 1:
        return candidates[0]

    parsed_wbs6 = _parsed_wbs6_code(parsed)
    if parsed_wbs6:
        normalized_wbs6 = _normalize_code_token(parsed_wbs6)
        filtered = [
            item
            for item in candidates
            if _normalize_code_token(item.wbs6_code) == normalized_wbs6
        ]
        if len(filtered) == 1:
            return filtered[0]
        if filtered:
            candidates = filtered

    return sorted(
        candidates,
        key=lambda item: (item.item_code or item.product_id or "").lower(),
    )[0]


def _parsed_wbs6_code(parsed: ParsedVoce) -> str | None:
    for level in parsed.wbs_levels:
        if getattr(level, "level", None) == 6:
            if level.code:
                return level.code
            if level.description:
                return level.description
    return None


def _description_signature(
    description: str | None,
    unit: str | None,
    wbs6_code: str | None,
) -> str | None:
    # We now rely exclusively on the extended description so we can match Excel
    # rows regardless of how codes, units or WBS labels were exported.
    token = _normalize_description_token(description)
    if not token:
        return None
    return token


def _description_signature_from_parsed(voce: ParsedVoce) -> str | None:
    wbs6_code = None
    for level in voce.wbs_levels:
        if level.level == 6:
            wbs6_code = level.code or level.description
            break
    return _description_signature(voce.descrizione, voce.unita_misura, wbs6_code)


def _description_signature_from_model(voce: VoceComputo) -> str | None:
    return _description_signature(voce.descrizione, voce.unita_misura, voce.wbs_6_code)


_SEMANTIC_DEFAULT_BUCKET = "__all__"
_SEMANTIC_MIN_SCORE = 0.58
_HEAD_TAIL_WORD_LIMIT = 30
_WORD_TOKENIZER = re.compile(r"[A-Za-z0-9]+")


def _head_tail_signatures(
    description: str | None,
    limit: int = _HEAD_TAIL_WORD_LIMIT,
) -> tuple[str, str]:
    if not description:
        return "", ""
    tokens = _tokenize_words(description)
    if not tokens:
        return "", ""
    head_tokens = tokens[:limit]
    tail_tokens = tokens[-limit:] if len(tokens) > limit else tokens
    return " ".join(head_tokens), " ".join(tail_tokens)


def _tokenize_words(text: str) -> list[str]:
    normalized = unicodedata.normalize("NFKD", text)
    normalized = "".join(ch for ch in normalized if not unicodedata.combining(ch))
    normalized = normalized.lower()
    return _WORD_TOKENIZER.findall(normalized)


def _match_price_list_item_semantic(
    parsed: ParsedVoce,
    embedding_map: dict[str, list[tuple[PriceListItem, list[float]]]],
) -> PriceListItem | None:
    if not embedding_map:
        return None
    text = _compose_semantic_text_from_parsed(parsed)
    if not text:
        return None
    try:
        query_vector = semantic_embedding_service.embed_text(text)
    except RuntimeError:
        return None
    if not query_vector:
        return None

    normalized_wbs6 = _normalize_code_token(_parsed_wbs6_code(parsed)) or _SEMANTIC_DEFAULT_BUCKET
    candidate_buckets = [
        embedding_map.get(normalized_wbs6, []),
        embedding_map.get(_SEMANTIC_DEFAULT_BUCKET, []),
    ]

    best_score = _SEMANTIC_MIN_SCORE
    best_item: PriceListItem | None = None
    seen_ids: set[int] = set()
    for bucket in candidate_buckets:
        if not bucket:
            continue
        for item, vector in bucket:
            if item.id in seen_ids:
                continue
            seen_ids.add(item.id)
            if not vector or len(vector) != len(query_vector):
                continue
            try:
                score = float(sum(float(a) * float(b) for a, b in zip(query_vector, vector)))
            except (TypeError, ValueError):
                continue
            if score > best_score:
                best_score = score
                best_item = item
    return best_item




def _build_matching_report(
    legacy_pairs: Sequence[tuple[VoceComputo | None, ParsedVoce]],
    excel_only_labels: Sequence[str] | None = None,
    excel_only_groups: Sequence[str] | None = None,
    quantity_mismatches: Sequence[str] | None = None,
    quantity_totals: dict[str, float] | None = None,
) -> dict[str, Any]:
    matched: list[dict[str, Any]] = []
    missing: list[dict[str, Any]] = []
    for legacy, parsed in legacy_pairs:
        project_label = _voce_label(legacy) if legacy else None
        excel_label = parsed.descrizione or parsed.codice or "Voce senza descrizione"
        entry = {
            "project_label": project_label,
            "excel_label": excel_label,
            "price": parsed.prezzo_unitario,
            "quantity": parsed.quantita,
        }
        metadata = parsed.metadata or {}
        if metadata.get("missing_from_return"):
            missing.append(entry)
        else:
            matched.append(entry)
    report = {
        "matched": matched,
        "missing": missing,
        "excel_only": list(excel_only_labels or []),
    }
    if excel_only_groups:
        report["excel_only_groups"] = list(excel_only_groups)
    if quantity_mismatches:
        report["quantity_mismatches"] = list(quantity_mismatches)
    if quantity_totals:
        report["quantity_totals"] = quantity_totals
        report["quantity_total_mismatch"] = (
            abs(quantity_totals.get("delta", 0.0)) > 1e-6
        )
    return report


def _describe_parsed_voce(voce: ParsedVoce) -> str:
    code = voce.codice or (f"PROG-{voce.progressivo}" if voce.progressivo is not None else "SENZA-CODICE")
    price = f"{voce.prezzo_unitario:.2f}" if isinstance(voce.prezzo_unitario, (int, float)) else "n/d"
    return f"{code} @ {price}"


def _log_unmatched_price_entries(entries: Sequence[ParsedVoce], limit: int = 5) -> None:
    samples = ", ".join(_describe_parsed_voce(voce) for voce in entries[:limit])
    logger.warning(
        "Import LC: %s righe non hanno trovato corrispondenza nel listino (prime: %s)",
        len(entries),
        samples or "n/d",
    )


def _log_price_conflicts(conflicts: Iterable[dict[str, Any]], limit: int = 5) -> None:
    conflicts = list(conflicts)
    if not conflicts:
        return
    formatted = []
    for entry in conflicts[:limit]:
        prices = entry.get("prices") or []
        formatted.append(
            f"{entry.get('item_code') or entry.get('price_list_item_id')} -> {sorted(prices)}"
        )
    logger.warning(
        "Import LC: %s codici con prezzi multipli (prime: %s)",
        len(conflicts),
        "; ".join(formatted) or "n/d",
    )


def _build_lc_matching_report(summary: dict[str, Any]) -> dict[str, Any]:
    price_items: Sequence[PriceListItem] = summary.get("price_items") or []
    matched_ids: set[int] = set(summary.get("matched_item_ids") or [])
    unmatched_entries: Sequence[ParsedVoce] = summary.get("unmatched_entries") or []

    missing_items = [
        {
            "price_list_item_id": item.id,
            "item_code": item.item_code,
            "item_description": item.item_description,
        }
        for item in price_items
        if item.id not in matched_ids
    ]

    unmatched_rows_sample = [
        _shorten_label(voce.descrizione or voce.codice or "voce senza descrizione")
        for voce in unmatched_entries[:10]
    ]

    return {
        "mode": "lc",
        "total_price_items": len(price_items),
        "matched_price_items": len(matched_ids),
        "missing_price_items": missing_items,
        "unmatched_rows_sample": unmatched_rows_sample,
        "price_conflicts": summary.get("conflicting_price_items") or [],
    }


def _build_project_description_buckets(
    progetto_voci: Sequence[VoceComputo],
) -> dict[str, list[tuple[VoceComputo, set[str]]]]:
    buckets: dict[str, list[tuple[VoceComputo, set[str]]]] = {}
    for voce in progetto_voci:
        key = _wbs_key_from_model(voce)
        if not key:
            continue
        tokens = _descr_tokens(voce.descrizione)
        buckets.setdefault(key, []).append((voce, tokens))
    return buckets


def _assign_wrapper_preferences(
    wrappers: Sequence[dict[str, Any]],
    project_buckets: dict[str, list[tuple[VoceComputo, set[str]]]],
) -> None:
    for wrapper in wrappers:
        base_key = wrapper.get("base_key")
        if not base_key:
            continue
        candidates = project_buckets.get(base_key)
        if not candidates:
            continue
        wrapper_tokens: set[str] = wrapper.get("tokens") or set()
        if not wrapper_tokens:
            continue
        best_voice: VoceComputo | None = None
        best_score = 0.0
        second_score = 0.0
        for voce, tokens in candidates:
            score = _jaccard_similarity(tokens, wrapper_tokens)
            if score > best_score:
                second_score = best_score
                best_score = score
                best_voice = voce
            elif score > second_score:
                second_score = score
        if best_voice and best_score >= 0.15 and (best_score - second_score) >= 0.01:
            wrapper["preferred_voice_id"] = best_voice.id


def _filter_entries_by_primary(
    entries: Sequence[tuple[ParsedVoce, VoceComputo]],
    primary: str | None,
) -> list[tuple[ParsedVoce, VoceComputo]]:
    if not primary:
        return list(entries)
    filtered: list[tuple[ParsedVoce, VoceComputo]] = []
    for parsed, legacy in entries:
        project_key = _wbs_key_from_model(legacy)
        project_primary, _ = _split_wbs_key(project_key)
        if project_primary == primary:
            filtered.append((parsed, legacy))
    return filtered


def _keys_from_voce_progetto(voce: VoceComputo) -> list[str]:
    keys: list[str] = []
    _append_description_tokens(keys, voce.descrizione)
    _append_token(keys, voce.wbs_7_description)
    _append_token(keys, voce.wbs_6_description)
    _append_token(keys, voce.wbs_5_description)
    _append_token(keys, voce.wbs_4_description)
    _append_token(keys, voce.wbs_3_description)
    _append_token(keys, voce.wbs_2_description)
    _append_token(keys, voce.wbs_1_description)
    _append_token(keys, voce.wbs_7_code)
    _append_token(keys, voce.wbs_6_code)
    _append_token(keys, voce.wbs_5_code)
    _append_token(keys, voce.wbs_4_code)
    _append_token(keys, voce.wbs_3_code)
    _append_token(keys, voce.wbs_2_code)
    _append_token(keys, voce.wbs_1_code)
    _append_token(keys, voce.codice)
    if voce.progressivo is not None:
        _append_token(keys, f"progressivo-{voce.progressivo}")
    return keys


def _keys_from_parsed_voce(voce: ParsedVoce) -> list[str]:
    keys: list[str] = []
    _append_description_tokens(keys, voce.descrizione)
    _append_token(keys, voce.codice)
    for livello in voce.wbs_levels:
        _append_token(keys, livello.description)
        _append_token(keys, livello.code)
    if voce.progressivo is not None:
        _append_token(keys, f"progressivo-{voce.progressivo}")
    return keys


def _append_token(target: list[str], value: str | None) -> None:
    token = _normalize_token(value)
    if token and token not in target:
        target.append(token)

def _descr_tokens(text: str | None) -> set[str]:
    if not text:
        return set()
    tokens: set[str] = set()
    # Parole comuni da ignorare (articoli, preposizioni, congiunzioni)
    stopwords = {
        "per", "con", "dei", "del", "dalla", "dallo", "dalle", "dagli",
        "alla", "allo", "alle", "agli", "nella", "nello", "nelle", "negli",
        "sulla", "sullo", "sulle", "sugli", "della", "dello", "delle", "degli",
        "una", "uno", "gli", "le", "il", "lo", "la", "di", "da", "in", "su",
        "a", "e", "o", "ma", "se", "che", "the", "of", "and", "or", "for",
        "with", "from", "to", "in", "on", "at", "by"
    }

    normalized_text = text.replace("\r\n", "\n").replace("\r", "\n")
    for segment in [normalized_text, *[part.strip() for part in normalized_text.split("\n") if part.strip()]]:
        token = _normalize_token(segment)
        if token and len(token) >= 6:
            tokens.add(token)

    for token in re.split(r"[^A-Za-z0-9]+", text):
        if len(token) >= 3 and token.lower() not in stopwords:
            tokens.add(token.lower())
    return tokens

def _append_description_tokens(target: list[str], value: str | None) -> None:
    """
    Tokenizza la descrizione in modo leggero:
    - intera descrizione / riga
    - singole parole con lunghezza >= 3 (abbassata da 4 a 3)
    Niente n-gram, altrimenti esplode il numero di chiavi.
    """
    if not value:
        return

    text = str(value).replace("\r\n", "\n").replace("\r", "\n")
    segments = [text]
    segments.extend(part.strip() for part in text.split("\n") if part.strip())

    # Stopwords da ignorare
    stopwords = {
        "per", "con", "dei", "del", "dalla", "dallo", "dalle", "dagli",
        "alla", "allo", "alle", "agli", "nella", "nello", "nelle", "negli",
        "sulla", "sullo", "sulle", "sugli", "della", "dello", "delle", "degli",
        "una", "uno", "gli", "le", "il", "lo", "la", "di", "da", "in", "su",
        "a", "e", "o", "ma", "se", "che"
    }

    for segment in segments:
        if not segment:
            continue
        # token "intera frase"
        _append_token(target, segment)

        # token parola singola (escluse stopwords)
        words = [
            w for w in re.split(r"[^A-Za-z0-9]+", segment)
            if len(w) >= 3 and w.lower() not in stopwords
        ]
        for word in words:
            _append_token(target, word)

def _normalize_token(value: str | None) -> str | None:
    if not value:
        return None
    normalized = unicodedata.normalize("NFKD", str(value))
    cleaned = "".join(ch.lower() for ch in normalized if ch.isalnum())
    return cleaned or None


def _pick_match(
    index: dict[str, list[dict[str, Any]]],
    keys: Sequence[str],
    voce_progetto: VoceComputo | None = None,
) -> ParsedVoce | None:
    """
    Versione ottimizzata: usa token pre-calcolati e limita il numero di candidati.
    """

    project_wbs_key = None
    project_base_key = None
    if voce_progetto:
        project_wbs_key = _wbs_key_from_model(voce_progetto)
        project_base_key = _base_wbs_key_from_key(project_wbs_key)
        if project_base_key:
            wbs_bucket = index.get(f"__wbs__:{project_base_key}")
            voce_wbs = _claim_wbs_bucket(wbs_bucket, voce_progetto)
            if voce_wbs:
                return voce_wbs

    if voce_progetto and voce_progetto.codice:
        token_code = _normalize_token(voce_progetto.codice)
        if token_code:
            bucket = index.get(token_code)
            if bucket:
                for wrapper in bucket:
                    base_scope = wrapper.get("base_key")
                    if base_scope and project_base_key and base_scope != project_base_key:
                        continue
                    preferred_id = wrapper.get("preferred_voice_id")
                    if (
                        preferred_id
                        and voce_progetto.id != preferred_id
                        and not wrapper.get("reusable")
                    ):
                        continue
                    if wrapper["used"]:
                        continue
                    voce_match = wrapper["voce"]
                    metadata = getattr(voce_match, "metadata", {}) or {}
                    if metadata.get("group_total_only"):
                        continue
                    wrapper["matched"] = True
                    wrapper["used"] = True
                    return voce_match

    # 2) Candidati da keys (ma evitando duplicati)
    candidates: list[dict[str, Any]] = []
    seen_wrappers: set[int] = set()
    for key in keys:
        if len(key) < 4:
            continue
        bucket = index.get(key)
        if not bucket:
            continue
        for wrapper in bucket:
            base_scope = wrapper.get("base_key")
            if base_scope and project_base_key and base_scope != project_base_key:
                continue
            preferred_id = wrapper.get("preferred_voice_id")
            if (
                preferred_id
                and voce_progetto
                and voce_progetto.id != preferred_id
                and not wrapper.get("reusable")
            ):
                continue
            wid = id(wrapper)
            if wid in seen_wrappers:
                continue
            if wrapper["used"]:
                continue
            seen_wrappers.add(wid)
            candidates.append(wrapper)
        # opzionale: hard cap per sicurezza
        if len(candidates) >= 100:
            break

    if not candidates:
        return None

    project_tokens = _descr_tokens(
        voce_progetto.descrizione if voce_progetto else None
    )
    if not project_tokens:
        return None

    # 3) Pre-filtra: almeno 1 parola forte in comune (abbassata da 2 a 1)
    filtered: list[tuple[dict[str, Any], set[str]]] = []
    for wrapper in candidates:
        ret_tokens = wrapper.get("tokens") or set()
        if len(project_tokens & ret_tokens) >= 1:
            filtered.append((wrapper, ret_tokens))

    if not filtered:
        return None

    # Limitiamo ulteriormente per sicurezza
    filtered = filtered[:30]

    # 4) Similarità Jaccard (soglia abbassata da 0.10 a 0.05)
    best_score = 0.0
    best_wrapper: dict[str, Any] | None = None
    for wrapper, ret_tokens in filtered:
        score = _jaccard_similarity(project_tokens, ret_tokens)
        if score > best_score:
            best_score = score
            best_wrapper = wrapper

    if not best_wrapper or best_score < 0.05:
        best_wrapper = _match_by_description_similarity(
            voce_progetto,
            [wrapper for wrapper in candidates if not wrapper.get("used")],
        )
        if not best_wrapper:
            return None

    voce = best_wrapper["voce"]
    best_wrapper["matched"] = True
    best_wrapper["used"] = True
    return voce


def _claim_wbs_bucket(
    bucket: list[dict[str, Any]] | None,
    voce_progetto: VoceComputo | None,
) -> ParsedVoce | None:
    if not bucket:
        return None
    project_tokens = _descr_tokens(voce_progetto.descrizione if voce_progetto else None)
    preferred_id = voce_progetto.id if voce_progetto else None
    best_wrapper: dict[str, Any] | None = None
    best_score = -1.0
    for wrapper in bucket:
        if wrapper["used"]:
            continue
        metadata = getattr(wrapper["voce"], "metadata", {}) or {}
        if metadata.get("group_total_only"):
            continue
        wrapper_preferred = wrapper.get("preferred_voice_id")
        if (
            preferred_id
            and wrapper_preferred
            and wrapper_preferred != preferred_id
            and not wrapper.get("reusable")
        ):
            continue
        if project_tokens and wrapper.get("tokens"):
            score = _jaccard_similarity(project_tokens, wrapper["tokens"])
        else:
            score = 0.0
        if score > best_score:
            best_score = score
            best_wrapper = wrapper
    if not best_wrapper:
        return None
    best_wrapper["matched"] = True
    base_scope = best_wrapper.get("base_key")
    if not (best_wrapper.get("reusable") and base_scope and voce_progetto):
        best_wrapper["used"] = True
    else:
        project_base = _base_wbs_key_from_key(_wbs_key_from_model(voce_progetto))
        if project_base != base_scope:
            best_wrapper["used"] = True
    voce = best_wrapper["voce"]
    if best_wrapper.get("reusable"):
        voce = copy.deepcopy(voce)
    return voce

def _voce_label(voce: VoceComputo) -> str:
    parts: list[str] = []
    if voce.progressivo is not None:
        parts.append(f"progressivo {voce.progressivo}")
    if voce.codice:
        parts.append(voce.codice)
    if voce.descrizione:
        parts.append(voce.descrizione)
    return " - ".join(parts) or f"Voce ordine {voce.ordine}"


def _sum_project_quantities(voci: Sequence[VoceComputo]) -> Decimal | None:
    total = Decimal("0")
    count = 0
    for voce in voci:
        if voce.quantita is None:
            continue
        total += Decimal(str(voce.quantita))
        count += 1
    return total if count else None


def _format_quantity_value(value: Decimal) -> str:
    text = format(value, ".4f")
    if "." in text:
        text = text.rstrip("0").rstrip(".")
    return text or "0"


def _progress_price_key(voce: VoceComputo | None) -> tuple[int, str] | None:
    if voce is None or voce.progressivo is None:
        return None
    return (voce.progressivo, _normalize_code_token(voce.codice) or "")


def _has_progressivi(voci: Sequence[ParsedVoce]) -> bool:
    return any(getattr(voce, "progressivo", None) is not None for voce in voci)


def _quantities_match(
    project_value: float | None, offered_value: float | None, tolerance: float = 1e-4
) -> bool:
    if project_value in (None,) or offered_value in (None,):
        return True
    project_dec = Decimal(str(project_value))
    offered_dec = Decimal(str(offered_value))
    return abs(project_dec - offered_dec) <= Decimal(str(tolerance))


def _prices_match(
    first_value: float | None, second_value: float | None, tolerance: float = 1e-4
) -> bool:
    if first_value in (None,) or second_value in (None,):
        return True
    first_dec = Decimal(str(first_value))
    second_dec = Decimal(str(second_value))
    return abs(first_dec - second_dec) <= Decimal(str(tolerance))


def _shorten_label(label: str, limit: int = 120) -> str:
    text = label.strip()
    if len(text) <= limit:
        return text
    return text[: limit - 3] + "..."


def _match_by_description_similarity(
    voce_progetto: VoceComputo | None,
    candidates: Sequence[dict[str, Any]],
    *,
    min_ratio: float = 0.30,  # Abbassata da 0.45 a 0.30 per matching più permissivo
) -> dict[str, Any] | None:
    if not voce_progetto or not candidates:
        return None
    target_tokens = _descr_tokens(voce_progetto.descrizione)
    if not target_tokens:
        return None
    best_ratio = 0.0
    best_wrapper: dict[str, Any] | None = None
    for wrapper in candidates:
        if wrapper.get("used"):
            continue
        voce = wrapper["voce"]
        metadata = getattr(voce, "metadata", {}) or {}
        if metadata.get("group_total_only"):
            continue
        candidate_tokens = wrapper.get("tokens") or _descr_tokens(voce.descrizione)
        if not candidate_tokens:
            continue
        overlap = target_tokens & candidate_tokens
        denom = max(len(target_tokens), len(candidate_tokens))
        if denom == 0:
            continue
        ratio = len(overlap) / denom
        if ratio > best_ratio:
            best_ratio = ratio
            best_wrapper = wrapper
    if not best_wrapper or best_ratio < min_ratio:
        return None
    return best_wrapper


def _match_excel_entry_fuzzy(
    voce_progetto: VoceComputo,
    excel_entries: Sequence[dict[str, Any]],
    candidate_indices: list[int],
    *,
    min_ratio: float = 0.30,
) -> int | None:
    if not candidate_indices:
        return None
    tokens = _descr_tokens(voce_progetto.descrizione)
    if not tokens:
        return None
    best_ratio = 0.0
    best_idx: int | None = None
    for idx in candidate_indices:
        entry_tokens = excel_entries[idx]["tokens"]
        if not entry_tokens:
            continue
        overlap = tokens & entry_tokens
        denom = len(tokens | entry_tokens)
        if denom == 0:
            continue
        ratio = len(overlap) / denom
        if ratio > best_ratio:
            best_ratio = ratio
            best_idx = idx
    if best_idx is None or best_ratio < min_ratio:
        return None
    return best_idx


_FORCED_ZERO_CODE_PREFIXES = ("A004010",)
_FORCED_ZERO_DESCRIPTION_KEYWORDS = (
    "mark up fee",
    "mark-up fee",
    "markup fee",
)


def _detect_forced_zero_violations(voci: Sequence[ParsedVoce]) -> list[str]:
    alerts: list[str] = []
    for voce in voci:
        if not _is_forced_zero_voce(voce):
            continue
        fields: list[str] = []
        if _is_nonzero(voce.quantita):
            fields.append(f"Q={_format_quantity_for_warning(voce.quantita)}")
        if _is_nonzero(voce.prezzo_unitario):
            fields.append(f"P={format(voce.prezzo_unitario, '.2f')}")
        if _is_nonzero(voce.importo):
            fields.append(f"I={format(voce.importo, '.2f')}")
        if fields:
            label = voce.codice or (voce.descrizione or "voce senza descrizione")
            alerts.append(f"{_shorten_label(label)} ({', '.join(fields)})")
    return alerts


def _is_nonzero(value: float | None, tolerance: float = 1e-6) -> bool:
    return value is not None and abs(value) > tolerance


def _is_forced_zero_voce(voce: ParsedVoce) -> bool:
    return _requires_zero_guard(voce.codice, voce.descrizione)


def _requires_zero_guard(code: str | None, description: str | None) -> bool:
    code_token = _normalize_code_token(code)
    if code_token:
        for prefix in _FORCED_ZERO_CODE_PREFIXES:
            if code_token.startswith(prefix):
                return True
    description_token = _normalize_description_token(description)
    if not description_token:
        return False
    return any(keyword in description_token for keyword in _FORCED_ZERO_DESCRIPTION_KEYWORDS)


def _build_zero_guard_entry(
    codice: str | None,
    descrizione: str | None,
    quantita: float | None,
    prezzo: float | None,
    importo: float | None,
) -> ParsedVoce:
    return ParsedVoce(
        ordine=0,
        progressivo=None,
        codice=codice,
        descrizione=descrizione,
        wbs_levels=[],
        unita_misura=None,
        quantita=quantita,
        prezzo_unitario=prezzo,
        importo=importo,
        note=None,
        metadata=None,
    )


def _normalize_code_token(code: str | None) -> str:
    if not code:
        return ""
    normalized = str(code).upper()
    return re.sub(r"[^A-Z0-9]", "", normalized)


def _normalize_description_token(text: str | None) -> str:
    if not text:
        return ""
    normalized = unicodedata.normalize("NFKD", text)
    normalized = "".join(ch for ch in normalized if not unicodedata.combining(ch))
    normalized = normalized.lower()
    normalized = re.sub(r"\s+", " ", normalized).strip()
    return normalized


def _format_quantity_for_warning(value: float) -> str:
    text = format(value, ".4f")
    if "." in text:
        text = text.rstrip("0").rstrip(".")
    return text or "0"


def _jaccard_similarity(a: set[str], b: set[str]) -> float:
    if not a or not b:
        return 0.0
    return len(a & b) / len(a | b)


def _price_bundle(voce: ParsedVoce) -> tuple[float | None, float | None, float | None]:
    prezzo = voce.prezzo_unitario
    quantita = voce.quantita
    importo = voce.importo
    return (
        round(prezzo, 4) if prezzo is not None else None,
        quantita,
        _ceil_amount(importo) if importo is not None else None,
    )


def _stabilize_return_price(
    value: float,
    reference_price: float | None,
) -> tuple[float, bool]:
    if reference_price in (None, 0):
        return value, False

    safe_reference = abs(reference_price)
    if safe_reference < 1e-6:
        return value, False

    # NEW: se il prezzo di riferimento è troppo basso (tipico delle voci a corpo
    # con Q.t = 0), non tentiamo di normalizzare: il valore di ritorno è più
    # affidabile del progetto.
    if safe_reference < 1:
        return value, False

    adjusted = value
    ratio = abs(adjusted) / safe_reference
    if ratio <= 250 or abs(adjusted) < 1000:
        return value, False

    for _ in range(4):
        adjusted /= 1000
        ratio = abs(adjusted) / safe_reference
        if ratio <= 250 or abs(adjusted) < 1000:
            return adjusted, True

    return value, False


def _wbs_key_from_model(voce: VoceComputo) -> str | None:
    primary = None
    for value in (
        voce.wbs_6_code,
        voce.wbs_6_description,
        voce.wbs_5_code,
        voce.wbs_5_description,
    ):
        token = _normalize_token(value)
        if token:
            primary = token
            break

    secondary = None
    for value in (
        voce.wbs_7_code,
        voce.wbs_7_description,
        voce.descrizione,
    ):
        token = _normalize_token(value)
        if token:
            secondary = token
            break

    if primary and secondary:
        return f"{primary}|{secondary}"
    return secondary or primary


def _wbs_key_from_parsed(voce: ParsedVoce) -> str | None:
    primary = None
    secondary = None
    description_token = _normalize_token(voce.descrizione)
    for livello in voce.wbs_levels:
        if livello.level == 6 and primary is None:
            primary = _normalize_token(livello.code) or _normalize_token(
                livello.description
            )
        if livello.level == 7 and secondary is None:
            secondary = _normalize_token(livello.code) or _normalize_token(
                livello.description
            )
    if secondary is None:
        secondary = (
            _normalize_token(voce.codice)
            or _normalize_token(voce.descrizione)
        )
    if primary and secondary:
        if description_token:
            return f"{primary}|{secondary}|{description_token}"
        return f"{primary}|{secondary}"
    if secondary and description_token and secondary != description_token:
        return f"{secondary}|{description_token}"
    return description_token or secondary or primary


def _wbs_base_key_from_parsed(voce: ParsedVoce) -> str | None:
    primary = None
    secondary = None
    for livello in voce.wbs_levels:
        if livello.level == 6 and primary is None:
            primary = _normalize_token(livello.code) or _normalize_token(
                livello.description
            )
        if livello.level == 7 and secondary is None:
            secondary = _normalize_token(livello.code) or _normalize_token(
                livello.description
            )
    if secondary is None:
        secondary = (
            _normalize_token(voce.codice)
            or _normalize_token(voce.descrizione)
        )
    if primary and secondary:
        return f"{primary}|{secondary}"
    return secondary or primary


def _split_wbs_key(key: str | None) -> tuple[str | None, str | None]:
    if not key:
        return None, None
    if "|" in key:
        primary, secondary = key.split("|", 1)
        return (primary or None), (secondary or None)
    return None, key


def _base_wbs_key_from_key(key: str | None) -> str | None:
    primary, secondary = _split_wbs_key(key)
    if primary and secondary:
        if "|" in secondary:
            secondary = secondary.split("|", 1)[0]
        return f"{primary}|{secondary}"
    if primary:
        return primary
    return secondary


def _collect_code_tokens(code: str | None) -> set[str]:
    if not code:
        return set()
    normalized = _normalize_token(code)
    tokens = set()
    if not normalized:
        return tokens
    tokens.add(normalized)
    segments = [segment for segment in re.split(r"[^A-Za-z0-9]+", code) if segment]
    builder = ""
    for segment in segments:
        cleaned = _normalize_token(segment)
        if not cleaned:
            continue
        builder += cleaned
        tokens.add(builder)
    return tokens


def _collect_description_tokens(text: str | None) -> set[str]:
    if not text:
        return set()
    tokens: set[str] = set()

    # Prima aggiungiamo il testo completo normalizzato (e le sue varianti per riga)
    normalized_text = text.replace("\r\n", "\n").replace("\r", "\n")
    for segment in [normalized_text, *[part.strip() for part in normalized_text.split("\n") if part.strip()]]:
        token = _normalize_token(segment)
        if token and len(token) >= 6:
            tokens.add(token)

    # Poi aggiungiamo i singoli pezzi (come prima)
    for segment in re.split(r"[^A-Za-z0-9]+", text):
        token = _normalize_token(segment)
        if token and len(token) >= 4:
            tokens.add(token)
    return tokens


def _distribute_group_targets(
    excel_targets: dict[str, Decimal],
    matched_entries: dict[str, list[ParsedVoce]],
    project_groups: dict[str, list[tuple[ParsedVoce, VoceComputo]]],
    project_code_groups: dict[str, list[tuple[ParsedVoce, VoceComputo]]],
    project_primary_groups: dict[str, list[tuple[ParsedVoce, VoceComputo]]],
    project_description_groups: dict[str, list[tuple[ParsedVoce, VoceComputo]]],
    excel_details: dict[str, dict[str, Any]],
) -> tuple[int, set[str]]:
    resolved = 0
    satisfied_keys: set[str] = set()
    for key, target in excel_targets.items():
        details = excel_details.get(key, {})
        project_entries: list[tuple[ParsedVoce, VoceComputo]] | None = None

        primary = details.get("primary")
        if primary is None:
            primary, _ = _split_wbs_key(key)

        description_tokens = list(details.get("description_tokens") or [])
        for token in description_tokens:
            candidates = project_description_groups.get(token)
            if not candidates:
                continue
            filtered = _filter_entries_by_primary(candidates, primary)
            if filtered:
                project_entries = filtered
                break

        if not project_entries:
            code_tokens = list(details.get("code_tokens") or [])
            for token in code_tokens:
                candidates = project_code_groups.get(token)
                if not candidates:
                    continue
                filtered = _filter_entries_by_primary(candidates, primary)
                if filtered:
                    project_entries = filtered
                    break

        if not project_entries:
            base_key = _base_wbs_key_from_key(key)
            if base_key:
                project_entries = project_groups.get(base_key)
            if not project_entries:
                project_entries = project_groups.get(key)

        if not project_entries and primary:
            project_entries = project_primary_groups.get(primary)
        if not project_entries:
            continue
        matched_list = matched_entries.get(key, [])
        matched_ids = {id(entry) for entry in matched_list}
        unmatched_entries = [
            (parsed, legacy)
            for parsed, legacy in project_entries
            if id(parsed) not in matched_ids
            and (parsed.metadata or {}).get("missing_from_return")
        ]
        if not unmatched_entries:
            continue
        assigned_total = sum(
            Decimal(str(entry.importo))
            for entry in matched_list
            if entry.importo is not None
        )
        residual = (target - assigned_total).quantize(
            Decimal("0.01"), rounding=ROUND_HALF_UP
        )
        if residual <= Decimal("0"):
            continue
        weight_values: list[Decimal] = []
        for _, legacy in unmatched_entries:
            base_value = legacy.importo
            if base_value is None and (
                legacy.quantita not in (None, 0)
                and legacy.prezzo_unitario is not None
            ):
                base_value = legacy.quantita * legacy.prezzo_unitario
            if base_value in (None, 0):
                weight_values.append(Decimal("0"))
            else:
                weight_values.append(Decimal(str(base_value)))
        total_weight = sum(weight_values)
        if total_weight <= Decimal("0"):
            weight_values = [Decimal("1")] * len(unmatched_entries)
            total_weight = Decimal(len(unmatched_entries))
        distributed = Decimal("0")
        allocated_any = False
        for index, ((parsed, legacy), weight) in enumerate(
            zip(unmatched_entries, weight_values)
        ):
            if index == len(unmatched_entries) - 1:
                share = residual - distributed
            else:
                share = (residual * weight / total_weight).quantize(
                    Decimal("0.01"), rounding=ROUND_HALF_UP
                )
                distributed += share
            if share <= Decimal("0"):
                continue
            parsed.quantita = legacy.quantita
            parsed.importo = float(share)
            if legacy.quantita not in (None, 0):
                price = (share / Decimal(str(legacy.quantita))).quantize(
                    Decimal("0.0001"), rounding=ROUND_HALF_UP
                )
                parsed.prezzo_unitario = float(price)
            elif legacy.prezzo_unitario is not None:
                parsed.prezzo_unitario = legacy.prezzo_unitario
            metadata = dict(parsed.metadata or {})
            metadata.pop("missing_from_return", None)
            metadata["group_allocation"] = {
                "wbs_key": key,
                "allocated_value": float(share),
            }
            parsed.metadata = metadata
            matched_entries.setdefault(key, []).append(parsed)
            resolved += 1
            allocated_any = True
        if allocated_any:
            satisfied_keys.add(key)
    return resolved, satisfied_keys


def _apply_rounding_to_match(
    entries: Sequence[ParsedVoce],
    target_total: Decimal,
) -> None:
    if not entries:
        return

    target_total = target_total.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

    import_values: list[Decimal] = []
    quantities: list[Decimal | None] = []
    adjustable: list[tuple[Decimal, int]] = []
    fallback: list[tuple[Decimal, int]] = []

    for idx, voce in enumerate(entries):
        quantita_val = voce.quantita
        prezzo_val = voce.prezzo_unitario
        importo_val = voce.importo
        preserve_price = (voce.metadata or {}).get("lock_return_price")

        quant_dec: Decimal | None = None
        if quantita_val not in (None, 0):
            quant_dec = Decimal(str(quantita_val))

        price_dec: Decimal | None = None
        if prezzo_val is not None:
            price_dec = Decimal(str(prezzo_val)).quantize(
                Decimal("0.0001"), rounding=ROUND_HALF_UP
            )

        if quant_dec is not None and price_dec is not None:
            import_dec = (price_dec * quant_dec).quantize(
                Decimal("0.01"), rounding=ROUND_HALF_UP
            )
            voce.prezzo_unitario = float(price_dec)
            voce.importo = float(import_dec)
            if not preserve_price:
                adjustable.append((import_dec.copy_abs(), idx))
        elif importo_val is not None:
            import_dec = Decimal(str(importo_val)).quantize(
                Decimal("0.01"), rounding=ROUND_HALF_UP
            )
            voce.importo = float(import_dec)
            if quant_dec is not None and not preserve_price:
                fallback.append((import_dec.copy_abs(), idx))
        else:
            import_dec = Decimal("0")

        import_values.append(import_dec)
        quantities.append(quant_dec)
        if not preserve_price and quant_dec is None and import_dec != Decimal("0"):
            fallback.append((import_dec.copy_abs(), idx))

    current_total = sum(import_values, Decimal("0"))
    difference = (target_total - current_total).quantize(
        Decimal("0.01"), rounding=ROUND_HALF_UP
    )
    if difference == Decimal("0"):
        return

    candidates = adjustable if adjustable else fallback
    if not candidates:
        return

    candidates.sort(key=lambda item: (item[0], item[1]), reverse=True)
    target_idx = candidates[0][1]

    new_import = (import_values[target_idx] + difference).quantize(
        Decimal("0.01"), rounding=ROUND_HALF_UP
    )
    if new_import < Decimal("0"):
        new_import = Decimal("0")
    import_values[target_idx] = new_import

    voce = entries[target_idx]
    quant_dec = quantities[target_idx]
    if quant_dec not in (None, Decimal("0")):
        new_price = (new_import / quant_dec).quantize(
            Decimal("0.0001"), rounding=ROUND_HALF_UP
        )
        voce.prezzo_unitario = float(new_price)
    voce.importo = float(new_import)


def _ceil_decimal_value(value: float | Decimal | int, exponent: str) -> Decimal:
    decimal_value = Decimal(str(value))
    return decimal_value.quantize(Decimal(exponent), rounding=ROUND_UP)


def _ceil_quantity(value: float | Decimal | int | None) -> float | None:
    if value is None:
        return None
    return float(_ceil_decimal_value(value, "0.000001"))


def _ceil_amount(value: float | Decimal | int | None) -> float | None:
    if value is None:
        return None
    return float(_ceil_decimal_value(value, "0.01"))

import_service = ImportService()
