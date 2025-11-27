#!/usr/bin/env python
"""
Script di test semplificato per debug import
"""
import sys
from pathlib import Path

# Aggiungi backend al path
sys.path.insert(0, str(Path(__file__).parent / "backend"))

# Import diretto dei moduli necessari evitando circular imports
from sqlmodel import Session, select, create_engine
from app.db.models import Commessa, Computo, ComputoTipo, PriceListItem, VoceComputo
from app.core.config import Settings

# Setup engine
settings = Settings()
database_path = settings.storage_root / settings.database_path
database_url = f"sqlite:///{database_path}"
engine = create_engine(database_url, echo=False)

def analyze_excel_structure(file_path: Path):
    """Analizza la struttura del file Excel"""
    from openpyxl import load_workbook

    print(f"\n📄 ANALISI FILE EXCEL: {file_path.name}")
    print("="*80)

    wb = load_workbook(filename=file_path, data_only=True, read_only=True)

    print(f"Fogli disponibili: {wb.sheetnames}")

    sheet = wb.active
    print(f"\nFoglio attivo: {sheet.title}")

    # Trova header
    header_row = None
    for i, row in enumerate(sheet.iter_rows(max_row=20, values_only=True), 1):
        if any(cell for cell in row):
            print(f"\nRiga {i}: {[str(cell)[:30] if cell else '' for cell in row[:10]]}")
            if any(str(cell).upper() in ["N.", "CODICE", "INDICAZIONE", "PU GARC", "QUANTITA'"] for cell in row if cell):
                header_row = i
                print(f"   ← POSSIBILE HEADER")
                break

    if header_row:
        headers = list(sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]
        print(f"\n✅ HEADER IDENTIFICATO (riga {header_row}):")
        for idx, h in enumerate(headers):
            if h:
                print(f"   Colonna {idx+1} ({chr(65+idx)}): {h}")

        # Mostra prime 5 righe dati
        print(f"\n📊 Prime 5 righe di dati:")
        data_rows = list(sheet.iter_rows(min_row=header_row+1, max_row=header_row+6, values_only=True))
        for i, row in enumerate(data_rows, 1):
            print(f"   Riga {header_row+i}: {[str(cell)[:20] if cell else 'NULL' for cell in row[:6]]}")

    wb.close()
    return header_row

def analyze_database_state(commessa_id: int):
    """Analizza lo stato del database per la commessa"""
    session = Session(engine)

    try:
        print(f"\n💾 ANALISI DATABASE (Commessa {commessa_id})")
        print("="*80)

        # Commessa
        commessa = session.get(Commessa, commessa_id)
        if not commessa:
            print(f"❌ Commessa {commessa_id} non trovata!")
            return

        print(f"Commessa: {commessa.nome}")

        # Elenco prezzi
        price_items = session.exec(
            select(PriceListItem).where(PriceListItem.commessa_id == commessa_id)
        ).all()
        print(f"\n📋 Elenco prezzi: {len(price_items)} voci")

        if not price_items:
            print("   ⚠️ NESSUN ELENCO PREZZI! Import LC non funzionerà.")
        else:
            print(f"   Prime 3 voci:")
            for item in price_items[:3]:
                code = item.item_code or 'NO CODE'
                desc = (item.item_description[:40] + '...') if item.item_description and len(item.item_description) > 40 else (item.item_description or 'NO DESC')
                print(f"      - {code} | {desc}")

        # Computo progetto
        computo_prog = session.exec(
            select(Computo)
            .where(
                Computo.commessa_id == commessa_id,
                Computo.tipo == ComputoTipo.progetto,
            )
            .order_by(Computo.created_at.desc())
        ).first()

        if not computo_prog:
            print(f"\n📐 Computo progetto: ❌ NON TROVATO")
            print("   ⚠️ Necessario per import MC/ritorno!")
            return

        print(f"\n📐 Computo progetto:")
        print(f"   ID: {computo_prog.id}")
        print(f"   Nome: {computo_prog.nome}")

        voci_prog = session.exec(
            select(VoceComputo)
            .where(VoceComputo.computo_id == computo_prog.id)
        ).all()

        print(f"   Voci totali: {len(voci_prog)}")

        with_prog = sum(1 for v in voci_prog if v.progressivo)
        print(f"   Voci con progressivo: {with_prog} ({100*with_prog/len(voci_prog) if voci_prog else 0:.1f}%)")

        with_price = sum(1 for v in voci_prog if v.prezzo_unitario and v.prezzo_unitario > 0)
        print(f"   Voci con prezzo: {with_price} ({100*with_price/len(voci_prog) if voci_prog else 0:.1f}%)")

        # Mostra sample
        if voci_prog:
            print(f"\n   Sample 3 voci con progressivo:")
            sample = [v for v in voci_prog if v.progressivo][:3]
            for voce in sample:
                prog = voce.progressivo or "N/A"
                code = voce.codice or "NO CODE"
                desc = (voce.descrizione[:40] + '...') if voce.descrizione and len(voce.descrizione) > 40 else (voce.descrizione or 'NO DESC')
                price = f"€{voce.prezzo_unitario:.2f}" if voce.prezzo_unitario else "NO PRICE"
                print(f"      #{prog} | {code} | {desc} | {price}")

        # Computi ritorno esistenti
        computi_ritorno = session.exec(
            select(Computo)
            .where(
                Computo.commessa_id == commessa_id,
                Computo.tipo == ComputoTipo.ritorno,
            )
        ).all()

        print(f"\n📦 Computi ritorno esistenti: {len(computi_ritorno)}")
        for comp in computi_ritorno:
            print(f"   - ID {comp.id} | {comp.impresa} | Round {comp.round_number} | €{comp.importo_totale or 0:,.2f}")
            if comp.note:
                print(f"      Note: {comp.note[:100]}")

    finally:
        session.close()

def main():
    COMMESSA_ID = 8
    FILE_PATH = Path(r"c:\Users\f.biggi\Taboolo\backend\storage\commessa_0008\uploads\20251127T150324_3600_ES_E_EC_02b_-_Computo_metrico_estimativo_Opere_civili.xlsx")

    print("\n🔍 DEBUG IMPORT - ANALISI PRELIMINARE")
    print("="*80)
    print(f"Commessa ID: {COMMESSA_ID}")
    print(f"File: {FILE_PATH.name}")

    if not FILE_PATH.exists():
        print(f"\n❌ File non trovato: {FILE_PATH}")
        return

    # Analisi Excel
    analyze_excel_structure(FILE_PATH)

    # Analisi DB
    analyze_database_state(COMMESSA_ID)

    print("\n" + "="*80)
    print("✅ Analisi completata")
    print("\nProssimi passi:")
    print("1. Verifica che le colonne Excel corrispondano a quelle configurate")
    print("2. Controlla che esistano voci nel computo progetto con progressivi")
    print("3. Per modalità LC: verifica che esista un elenco prezzi")
    print("="*80)

if __name__ == "__main__":
    main()
